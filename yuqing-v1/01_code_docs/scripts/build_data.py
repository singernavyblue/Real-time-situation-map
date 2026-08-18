# -*- coding: utf-8 -*-
"""从统计数据 Excel 生成态势大屏数据文件 data.js

支持两种数据来源：
1. 默认：扫描 16 个报表式 Excel（第一阶段原有流程）；
2. --atomic / ATOMIC_XLSX：读取“报表 → 数据库”原子化标准表（舆情事实表等），
   输出与第一种完全同结构的 data.js。
"""
import os, re, sys, json, datetime
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SUMMARY_NAME = "民族团结进步促进法舆情统计最终版_非支持原话全量汇总最终版.xlsx"

def _resolve_root():
    env = os.environ.get("DASH_ROOT")
    if env and os.path.isdir(env):
        return env
    parent = os.path.abspath(os.path.join(_SCRIPT_DIR, ".."))
    for cand in (parent, os.path.join(parent, "data")):
        if os.path.exists(os.path.join(cand, SUMMARY_NAME)):
            return cand
    pkg = os.path.abspath(os.path.join(_SCRIPT_DIR, "..", "..", "02_data"))
    if os.path.isdir(pkg):
        return pkg
    return parent

def _resolve_out():
    env = os.environ.get("DASH_OUT")
    if env:
        return env
    pkg = os.path.abspath(os.path.join(_SCRIPT_DIR, "..", "dashboard", "assets"))
    if os.path.isdir(pkg):
        return pkg
    return os.path.join(_resolve_root(), "dashboard", "assets")

ROOT = _resolve_root()
OUT_DIR = _resolve_out()
os.makedirs(OUT_DIR, exist_ok=True)
print("数据目录:", ROOT)
print("输出目录:", OUT_DIR)

def load_sheet(path, sheet):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    return rows

def first_nonempty_row(rows):
    for r in rows:
        if any(c is not None and str(c).strip() != "" for c in r):
            return r
    return rows[0]

def col_index(header, names):
    for i, h in enumerate(header):
        hs = str(h).strip()
        for n in names:
            if n in hs:
                return i
    return -1

def parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        if 40000 < v < 60000:
            try:
                return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=v)).date()
            except Exception:
                return None
        return None
    s = str(v).strip()
    m = re.search(r"(\d{4})[-/年.](\d{1,2})[-/月.](\d{1,2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime.date(y, mo, d)
        except Exception:
            return None
    return None

def clean_text(s):
    if s is None:
        return ""
    t = str(s).replace("\u3000", " ").replace("\xa0", " ").strip()
    t = re.sub(r"\s+", " ", t)
    return t

def is_meaningful_quote(t):
    if len(t) < 4:
        return False
    if t in ("未获取", "无", "无文字"):
        return False
    if "未获取" in t and len(t) < 8:
        return False
    for bad in ("截图未完整展示", "未完整展示", "未逐字", "未取得", "待复核", "正向点赞表情", "平台分组", "原话/代表性原话", "来源账号/栏目", "态度类别"):
        if bad in t:
            return False
    # 去掉纯表情/符号
    emoji_like = re.sub(r"[\U0001F000-\U0001FAFF\u2600-\u27BF\uFE0F]", "", t)
    if len(emoji_like.strip()) < 2:
        return False
    if re.fullmatch(r"[👍🙏❤️🌹🎉🤝😁😂💪]*", t):
        return False
    return True

# ================= 原子化模式：读取“报表 → 数据库”改造后的标准表 =================
def _f(rec, keys, default=""):
    for k in keys:
        if k in rec and rec[k] is not None:
            return rec[k]
    return default

def _aint(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", "").replace("，", "")
    m = re.search(r"\d+", s)
    return int(m.group()) if m else 0

def _atomic_counted(rec):
    if "是否计入统计" not in rec:
        return True
    v = rec["是否计入统计"]
    return v is not None and str(v).strip() in ("是", "1", "TRUE", "True", "计入", "已计入")

def _bucket_of(att):
    s = str(att or "")
    if any(k in s for k in ("支持", "认可", "赞同", "肯定")):
        return "support"
    if any(k in s for k in ("中性", "观点不明")):
        return "neutral"
    if "建议" in s:
        return "suggest"
    if any(k in s for k in ("非支持", "批评", "投诉", "担忧", "咨询", "公平", "歧视", "实施", "维权", "质疑", "不了解")):
        return "non_support"
    return "other"

def _issue_of(att, issue):
    s = str(issue or "").strip()
    if s and s != "-":
        return s
    a = str(att or "")
    for name, keys in [
        ("咨询疑问", ("咨询", "疑问")),
        ("担忧影响", ("担忧",)),
        ("明确批评", ("批评", "质疑")),
        ("投诉维权", ("投诉", "维权")),
        ("实施问题", ("实施",)),
        ("公平争议", ("公平",)),
        ("歧视偏见", ("歧视",)),
        ("不了解该法律", ("不了解",)),
    ]:
        if any(k in a for k in keys):
            return name
    return ""

def _platform_group_of(name):
    s = str(name or "")
    rules = [
        ("微博/热榜", ("微博", "热榜")),
        ("抖音等视频平台", ("抖音", "火山", "视频平台")),
        ("快手", ("快手",)),
        ("小红书/豆瓣等平台", ("小红书", "豆瓣")),
        ("知乎/B站/百度知道", ("知乎", "哔哩哔哩", "B站", "百度知道")),
        ("贴吧/头条/新闻评论", ("贴吧", "头条", "新闻")),
        ("微信公众号/视频号", ("微信", "公众号", "视频号")),
    ]
    for g, keys in rules:
        if any(k in s for k in keys):
            return g
    return "其他"

def _detect_atomic_xlsx():
    want = "--atomic" in sys.argv or bool(os.environ.get("ATOMIC_XLSX"))
    if not want:
        return None
    p = os.environ.get("ATOMIC_XLSX", "").strip()
    if p and os.path.exists(p):
        return p
    args = sys.argv[1:]
    if "--atomic" in args:
        i = args.index("--atomic")
        if i + 1 < len(args) and not args[i + 1].startswith("--"):
            cand = args[i + 1]
            if os.path.exists(cand):
                return cand
    roots = [
        ROOT,
        os.path.abspath(os.path.join(_SCRIPT_DIR, "..", "..", "..", "outputs")),
        os.path.abspath(os.path.join(_SCRIPT_DIR, "..", "..", "03_live_system", "docs")),
    ]
    for root in roots:
        if not os.path.isdir(root):
            continue
        for dirpath, dirs, files in os.walk(root):
            for fn in files:
                if fn.lower().endswith(".xlsx") and ("原子化" in fn or ("数据库" in fn and "示例" in fn)):
                    return os.path.join(dirpath, fn)
            depth = dirpath[len(root):].count(os.sep)
            if depth >= 2:
                dirs[:] = []
    return None

def read_atomic_fact(path):
    rows = load_sheet(path, "舆情事实表")
    non_empty = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    hidx = None
    for i, r in enumerate(non_empty[:6]):
        joined = "|".join(str(c) for c in r if c is not None)
        if "唯一编号" in joined and "正文" in joined and "是否计入统计" in joined:
            hidx = i
            break
    if hidx is None:
        raise SystemExit("[atomic] 未在“舆情事实表”中找到表头（需含：唯一编号 / 正文 / 是否计入统计）")
    header = non_empty[hidx]
    idx = {}
    for i, h in enumerate(header):
        if h is not None and str(h).strip():
            idx.setdefault(str(h).strip(), i)
    out = []
    for r in non_empty[hidx + 1:]:
        rec = {n: (r[i] if i < len(r) else None) for n, i in idx.items()}
        if any(v is not None and str(v).strip() != "" for v in rec.values()):
            out.append(rec)
    return out

def compute_atomic_sources(fact_recs):
    """从舆情事实表统计监测来源：监测来源=各平台去重账号数，查看信息=各平台记录条数"""
    per_platform = {}
    for rec in fact_recs:
        plat = clean_text(_f(rec, ["平台"])) or "未分类"
        account = clean_text(_f(rec, ["账号"]))
        # 账号为空（如热榜话题）按“平台本身”计 1 个来源，避免漏计
        src_key = account or f"{plat}（无账号）"
        d = per_platform.setdefault(plat, {"accounts": set(), "records": 0})
        d["accounts"].add(src_key)
        d["records"] += 1
    rows = []
    for plat in sorted(per_platform):
        d = per_platform[plat]
        rows.append(["监测来源", "原平台组", plat, len(d["accounts"]),
                     "自动统计：事实表去重账号数（无账号记录按平台计1）"])
        rows.append(["查看信息", "原平台组", plat, d["records"],
                     "自动统计：事实表记录条数"])
    return rows


def write_atomic_sources(path, rows):
    """把自动统计结果回写「监测来源与查看信息」表（保留表头与样式）"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path)
        if "监测来源与查看信息" not in wb.sheetnames:
            return
        ws = wb["监测来源与查看信息"]
        header_row = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
            joined = "|".join(str(c) for c in row if c is not None)
            if "类别" in joined and "数量" in joined:
                header_row = i
                break
        if header_row is None:
            return
        data = [list(r) for r in rows]
        sources_total = sum(r[3] for r in data if r[0] == "监测来源")
        views_total = sum(r[3] for r in data if r[0] == "查看信息")
        data.append(["监测来源", "合计（自动）", "-", sources_total,
                     "自动统计：按平台去重账号数合计"])
        data.append(["查看信息", "合计（自动）", "-", views_total,
                     "自动统计：事实表记录条数合计"])
        # 清空旧数据区，再写入新行
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            for c in row:
                c.value = None
        for ri, vals in enumerate(data):
            excel_row = header_row + 1 + ri
            for ci, v in enumerate(vals, start=1):
                ws.cell(row=excel_row, column=ci, value=v)
        tmp = path + ".tmp"
        wb.save(tmp)
        os.replace(tmp, path)
        print(f"[atomic] 监测来源与查看信息已从事实表自动重算（{len(data)} 行）")
    except Exception as e:
        print("[atomic] 监测来源表回写失败（不影响 data.js）:", e)

def read_atomic_staging_count(path):
    rows = load_sheet(path, "待清洗区")
    non_empty = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    hidx = None
    for i, r in enumerate(non_empty[:6]):
        joined = "|".join(str(c) for c in r if c is not None)
        if "清洗状态" in joined or "未通过原因" in joined:
            hidx = i
            break
    if hidx is None:
        return 0
    header = non_empty[hidx]
    i_status = col_index(header, ["清洗状态"])
    return sum(
        1 for r in non_empty[hidx + 1:]
        if 0 <= i_status < len(r) and "待复核" in str(r[i_status] or "")
    )

def build_from_atomic(path):
    fact_recs = read_atomic_fact(path)
    counted = []
    for rec in fact_recs:
        text = _f(rec, ["正文"])
        if not text or not str(text).strip():
            continue
        if not _atomic_counted(rec):
            continue
        counted.append(rec)

    platforms = {}
    regions = {}
    provinces = {}
    langs = {}
    trend = {}
    nonsup = {}
    quotes = []
    hot = []
    total = support = neutral = suggest = non_support = other = 0
    region_total = region_support = 0
    minority = 0

    for rec in counted:
        text = clean_text(_f(rec, ["正文"]))
        att = clean_text(_f(rec, ["总体态度", "态度"]))
        issue = _issue_of(att, clean_text(_f(rec, ["具体问题类别"])))
        bucket = _bucket_of(att)
        plat = clean_text(_f(rec, ["平台"]))
        pgroup = clean_text(_f(rec, ["平台组"])) or _platform_group_of(plat) or "其他"
        region = clean_text(_f(rec, ["地区"]))
        rgroup = clean_text(_f(rec, ["地区组"])) or region or "未分组"
        _prov_fallback = region
        _placeholder_mark = ("全国", "境外", "国际", "海峡", "未获取", "县级市", "涉及地区", "行政层级")
        if _prov_fallback in ("自治州", "地区", "") or any(k in _prov_fallback for k in _placeholder_mark):
            _prov_fallback = ""
        province = clean_text(_f(rec, ["省份"])) or _prov_fallback
        lang = clean_text(_f(rec, ["语言"])) or "中文"
        date = parse_date(_f(rec, ["发布时间", "发布日期"]))
        likes = _aint(_f(rec, ["点赞量"]))
        comments = _aint(_f(rec, ["评论量"]))
        shares = _aint(_f(rec, ["转发量"]))
        account = clean_text(_f(rec, ["账号"]))
        source = clean_text(_f(rec, ["具体来源"])) or os.path.basename(path)
        mino_flag = str(_f(rec, ["是否少数民族语言"])).strip() in ("是", "1", "TRUE", "True")
        if not mino_flag:
            mino_flag = any(k in lang for k in ("彝", "藏", "维吾尔", "哈萨克", "蒙古", "苗", "壮", "回", "朝鲜", "满"))
        is_minority = 1 if mino_flag else 0

        total += 1
        if bucket == "support":
            support += 1
        elif bucket == "neutral":
            neutral += 1
        elif bucket == "suggest":
            suggest += 1
        elif bucket == "non_support":
            non_support += 1
            nonsup[issue or "其他/未分类"] = nonsup.get(issue or "其他/未分类", 0) + 1
        else:
            other += 1
        if region or province:
            region_total += 1
            if bucket == "support":
                region_support += 1
        if is_minority:
            minority += 1

        p = platforms.setdefault(pgroup, {
            "name": pgroup, "total": 0, "support": 0, "neutral": 0, "qa": 0, "worry": 0,
            "criticism": 0, "complaint": 0, "implement": 0, "fairness": 0,
            "discrimination": 0, "suggest": 0, "unknownLaw": 0, "other": 0,
            "supportRate": 0.0, "nonSupport": 0,
        })
        p["total"] += 1
        p["support"] += 1 if bucket == "support" else 0
        p["neutral"] += 1 if bucket == "neutral" else 0
        p["suggest"] += 1 if bucket == "suggest" else 0
        if bucket == "non_support":
            p["nonSupport"] += 1
            cat = issue
            if cat in ("咨询疑问", "担忧影响", "明确批评", "投诉维权", "实施问题", "公平争议", "歧视偏见", "不了解该法律"):
                p[{"咨询疑问": "qa", "担忧影响": "worry", "明确批评": "criticism", "投诉维权": "complaint",
                   "实施问题": "implement", "公平争议": "fairness", "歧视偏见": "discrimination",
                   "不了解该法律": "unknownLaw"}[cat]] += 1
            else:
                p["other"] += 1

        rg = regions.setdefault(rgroup, {
            "name": rgroup, "total": 0, "support": 0, "neutral": 0, "qa": 0, "worry": 0,
            "criticism": 0, "complaint": 0, "implement": 0, "fairness": 0,
            "discrimination": 0, "pending": 0,
        })
        rg["total"] += 1
        rg["support"] += 1 if bucket == "support" else 0
        rg["neutral"] += 1 if bucket == "neutral" else 0
        if bucket == "non_support":
            cat = issue
            if cat in ("咨询疑问", "担忧影响", "明确批评", "投诉维权", "实施问题", "公平争议", "歧视偏见"):
                rg[{"咨询疑问": "qa", "担忧影响": "worry", "明确批评": "criticism", "投诉维权": "complaint",
                    "实施问题": "implement", "公平争议": "fairness", "歧视偏见": "discrimination"}[cat]] += 1

        pv = provinces.setdefault(province, {
            "name": province, "short": province, "value": 0, "total": 0, "support": 0,
            "neutral": 0, "qa": 0, "worry": 0, "criticism": 0, "complaint": 0,
            "implement": 0, "fairness": 0, "discrimination": 0, "suggest": 0, "other": 0,
            "groups": [],
        })
        pv["value"] += 1
        pv["total"] += 1
        pv["support"] += 1 if bucket == "support" else 0
        pv["neutral"] += 1 if bucket == "neutral" else 0
        if bucket == "suggest":
            pv["suggest"] += 1
        if bucket == "non_support":
            cat = issue
            if cat in ("咨询疑问", "担忧影响", "明确批评", "投诉维权", "实施问题", "公平争议", "歧视偏见", "不了解该法律"):
                pv[{"咨询疑问": "qa", "担忧影响": "worry", "明确批评": "criticism", "投诉维权": "complaint",
                    "实施问题": "implement", "公平争议": "fairness", "歧视偏见": "discrimination",
                    "不了解该法律": "unknownLaw"}[cat]] += 1
            else:
                pv["other"] += 1
        elif bucket == "other":
            pv["other"] += 1
        if rgroup and rgroup not in [g["group"] for g in pv["groups"]]:
            pv["groups"].append({"group": rgroup, "total": regions[rgroup]["total"]})

        lv = langs.get(lang, {"name": lang, "value": 0})
        lv["value"] += 1
        langs[lang] = lv

        if date:
            trend[date.isoformat()] = trend.get(date.isoformat(), 0) + 1

        quotes.append({
            "platform": plat,
            "region": region or province,
            "group": rgroup,
            "attitude": att or "待核实",
            "text": text,
            "date": date.isoformat() if date else "",
            "source": source,
            "src": source,
            "account": account,
            "language": lang,
        })

        if likes > 0:
            hot.append({
                "platform": plat or pgroup,
                "account": account,
                "title": text[:60],
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "date": date.isoformat() if date else "",
            })

    platforms_list = sorted(platforms.values(), key=lambda x: x["total"], reverse=True)
    for p in platforms_list:
        p["supportRate"] = round(p["support"] / p["total"] * 100, 2) if p["total"] else 0.0
    provinces_list = sorted(
        [p for p in provinces.values() if p.get("name")],
        key=lambda x: x["total"],
        reverse=True,
    )
    for pv in provinces_list:
        for g in pv.get("groups", []):
            if g["group"] in regions:
                g["total"] = regions[g["group"]]["total"]
    # 地区轮播维度：省级（每个省份一条，地图高亮单个省）
    # 保留 sourceGroup（原地区组），让“地区信息流”也能匹配按地区组归类的原话
    regions_list = []
    for pv in provinces_list:
        groups = pv.get("groups") or []
        regions_list.append({
            "name": pv["name"],
            "total": pv["total"],
            "support": pv["support"],
            "neutral": pv["neutral"],
            "qa": pv["qa"],
            "worry": pv["worry"],
            "criticism": pv["criticism"],
            "complaint": pv["complaint"],
            "implement": pv["implement"],
            "fairness": pv["fairness"],
            "discrimination": pv["discrimination"],
            "pending": pv.get("pending", 0),
            "provinces": [pv["name"]],
            "sourceGroup": groups[0].get("group") if groups else "",
        })
    langs_list = sorted(langs.values(), key=lambda x: x["value"], reverse=True)
    hot.sort(key=lambda x: x["likes"], reverse=True)
    hot = hot[:9]
    quotes.sort(key=lambda q: (len(q["text"]), q["attitude"] == "支持认可"), reverse=True)
    trend_series = [{"date": d, "value": v} for d, v in sorted(trend.items())]

    source_rows = compute_atomic_sources(counted)
    sources = sum(r[3] for r in source_rows if r[0] == "监测来源")
    views = sum(r[3] for r in source_rows if r[0] == "查看信息")
    write_atomic_sources(path, source_rows)

    lang_spotlight = []
    for l in langs_list:
        if any(k in l["name"] for k in ("中文", "普通话", "汉语")):
            continue
        sample = next((r for r in counted if clean_text(_f(r, ["语言"])) == l["name"]), None)
        lang_spotlight.append({
            "lang": l["name"],
            "record": (clean_text(_f(sample, ["正文"]))[:60] if sample else "（原子事实表记录）"),
            "count": l["value"],
        })
    chinese_count = sum(l["value"] for l in langs_list if any(k in l["name"] for k in ("中文", "普通话", "汉语")))
    lang_spotlight.append({"lang": "中文/普通话", "record": "全网监测主体为中文评论（原子事实表重算）", "count": chinese_count})

    detail = [
        {"name": "支持认可", "value": support},
        {"name": "中性信息", "value": neutral},
        {"name": "咨询疑问", "value": nonsup.get("咨询疑问", 0)},
        {"name": "担忧影响", "value": nonsup.get("担忧影响", 0)},
        {"name": "明确批评", "value": nonsup.get("明确批评", 0)},
        {"name": "投诉维权", "value": nonsup.get("投诉维权", 0)},
        {"name": "实施问题", "value": nonsup.get("实施问题", 0)},
        {"name": "公平争议", "value": nonsup.get("公平争议", 0)},
        {"name": "歧视偏见", "value": nonsup.get("歧视偏见", 0)},
        {"name": "参与建议", "value": suggest},
        {"name": "不了解该法律", "value": nonsup.get("不了解该法律", 0)},
    ]

    data = {
        "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d"),
        "topStats": {
            "totalOpinions": total,
            "totalOpinionsLabel": "平台维度公众意见总量",
            "regionOpinions": region_total,
            "regionOpinionsLabel": "可明确归属地区的公众意见",
            "supportCount": support,
            "supportRate": round(support / total * 100, 2) if total else 0.0,
            "regionSupport": region_support,
            "regionSupportRate": round(region_support / region_total * 100, 2) if region_total else 0.0,
            "nonSupport": non_support,
            "nonSupportLabel": "非支持/非肯定态度",
            "monitorSources": sources,
            "monitorSourcesLabel": "监测来源数",
            "viewedInfo": views,
            "viewedInfoLabel": "查看信息数",
            "minorityLang": minority,
            "minorityLangLabel": "少数民族语言舆情",
        },
        "platforms": platforms_list,
        "languagePlatform": langs_list,
        "languageRegion": [],
        "langSpotlight": lang_spotlight,
        "regions": regions_list,
        "provinces": provinces_list,
        "attitude": {
            "macro": [
                {"name": "支持认可", "value": support},
                {"name": "中性信息", "value": neutral},
                {"name": "参与建议", "value": suggest},
                {"name": "非支持/非肯定", "value": non_support},
            ],
            "detail": detail,
        },
        "nonSupport": [{"name": k, "value": v} for k, v in nonsup.items()],
        "trend": trend_series,
        "quotes": quotes,
        "hotTop": hot,
    }
    return data

def write_outputs(data):
    # 原子替换：先写临时文件，再 os.replace，避免大屏读到写了一半的 data.js
    js_path = os.path.join(OUT_DIR, "data.js")
    tmp_js = js_path + ".tmp"
    with open(tmp_js, "w", encoding="utf-8") as f:
        f.write("window.DASH_DATA = ")
        json.dump(data, f, ensure_ascii=False, indent=1)
        f.write(";\n")
    os.replace(tmp_js, js_path)
    print("data.js 已生成（原子替换）:", js_path)

    # 地图：拆分为大陆主图 + 南海诸岛插图（避免 file:// 下 fetch CORS 问题）
    china_path = os.path.join(OUT_DIR, "china.json")
    if os.path.exists(china_path):
        with open(china_path, encoding="utf-8") as f:
            geo = json.load(f)
        main_features = []
        nanhai_feature = None
        nanhai_coords = []
        for feat in geo["features"]:
            name = feat["properties"].get("name") or ""
            if name == "" or feat["properties"].get("adchar") == "JD":
                feat["properties"]["name"] = "南海诸岛"
                nanhai_feature = feat
                nanhai_coords.extend(feat["geometry"]["coordinates"])
            elif name == "海南省":
                polys = feat["geometry"]["coordinates"]
                keep = []
                move = []
                for poly in polys:
                    rings = poly if feat["geometry"]["type"] == "MultiPolygon" and poly and isinstance(poly[0][0], list) else [poly]
                    miny = min(pt[1] for ring in rings for pt in ring)
                    if miny >= 17.5:
                        keep.append(poly)
                    else:
                        move.append(poly)
                if keep:
                    feat["geometry"]["coordinates"] = keep
                    main_features.append(feat)
                nanhai_coords.extend(move)
            else:
                main_features.append(feat)
        main_geo = {"type": "FeatureCollection", "features": main_features}
        if nanhai_feature is None:
            nanhai_feature = {"type": "Feature", "properties": {"name": "南海诸岛"}, "geometry": {"type": "MultiPolygon", "coordinates": []}}
        nanhai_feature["geometry"]["coordinates"] = nanhai_coords
        nanhai_geo = {"type": "FeatureCollection", "features": [nanhai_feature]}
        china_js = os.path.join(OUT_DIR, "china.js")
        tmp_china = china_js + ".tmp"
        with open(tmp_china, "w", encoding="utf-8") as f:
            f.write("window.CHINA_GEO = ")
            json.dump(main_geo, f, ensure_ascii=False)
            f.write(";\nwindow.NANHAI_GEO = ")
            json.dump(nanhai_geo, f, ensure_ascii=False)
            f.write(";\n")
        os.replace(tmp_china, china_js)
        print("china.js 已生成（原子替换）：主图要素", len(main_features), "，南海插图要素", 1)
    else:
        print("提示：未找到 china.json，跳过地图拆分")

ATOMIC_XLSX = _detect_atomic_xlsx()
if ATOMIC_XLSX:
    print("原子化模式（读取标准数据库表）:", ATOMIC_XLSX)
    data = build_from_atomic(ATOMIC_XLSX)
    write_outputs(data)
    print("平台:", len(data["platforms"]), "合计:", data["topStats"]["totalOpinions"])
    print("省级轮播:", len(data["regions"]), "合计:", sum(r["total"] for r in data["regions"]))
    print("省级行政区:", len(data["provinces"]))
    print("评论原话池:", len(data["quotes"]))
    print("趋势天数:", len(data["trend"]))
    print("高热内容:", len(data["hotTop"]))
    try:
        staging = read_atomic_staging_count(ATOMIC_XLSX)
        if staging:
            print("提示：待清洗区还有", staging, "条待复核记录未计入统计")
    except Exception:
        pass
    sys.exit(0)

# ---------------- 1. 汇总口径（以新“最终版_非支持原话全量汇总”为权威来源） ----------------
summary_file = os.path.join(ROOT, "民族团结进步促进法舆情统计最终版_非支持原话全量汇总最终版.xlsx")

def read_table_rows(path, sheet, header_names, skip_title_rows=1):
    rows = load_sheet(path, sheet)
    non_empty = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
    # 找表头行
    hidx = None
    for i, r in enumerate(non_empty):
        joined = " | ".join(str(c) for c in r if c is not None)
        if all(any(n in str(c) for c in r) for n in header_names):
            hidx = i
            break
    if hidx is None:
        for i, r in enumerate(non_empty[:6]):
            joined = "|".join(str(c) for c in r if c is not None)
            if "公众意见" in joined or "数量" in joined or "态度" in joined:
                hidx = i
                break
    if hidx is None:
        return []
    header = non_empty[hidx]
    # 记录所有有表头的列，避免遗漏态度列
    idx = {}
    for i, h in enumerate(header):
        if h is not None and str(h).strip():
            idx.setdefault(str(h).strip(), i)
    out = []
    for r in non_empty[hidx+1:]:
        rec = {}
        for n, i in idx.items():
            rec[n] = r[i] if i < len(r) else None
        if any(v is not None and str(v).strip() != "" for v in rec.values()):
            # 跳过表尾说明行：主数量列必须为数值
            primary = next((i for n, i in idx.items() if n != header_names[0]), -1)
            if primary >= 0 and len(r) > primary and not isinstance(r[primary], (int, float)):
                continue
            out.append(rec)
    return out

def to_int(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", "").strip()
    m = re.search(r"\d+", s)
    return int(m.group()) if m else 0

def to_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("%", "").strip()
    try:
        return float(s)
    except Exception:
        m = re.search(r"[\d.]+", s)
        return float(m.group()) if m else 0.0

def pick(rec, names):
    for n in names:
        if n in rec and rec[n] is not None:
            return rec[n]
    return None

# 平台统计
platform_rows = read_table_rows(summary_file, "平台统计", ["平台", "有效公众意见总数"])
platforms = []
for r in platform_rows:
    name = clean_text(r.get("平台"))
    if not name or name == "合计":
        continue
    platforms.append({
        "name": name,
        "total": to_int(r.get("有效公众意见总数")),
        "support": to_int(r.get("支持认可")),
        "neutral": to_int(pick(r, ["中性信息", "中性/观点不明"])),
        "qa": to_int(r.get("咨询疑问")),
        "worry": to_int(r.get("担忧影响")),
        "criticism": to_int(r.get("明确批评")),
        "complaint": to_int(pick(r, ["投诉维权", "投诉维权/举报"])),
        "implement": to_int(r.get("实施问题")),
        "fairness": to_int(r.get("公平争议")),
        "discrimination": to_int(r.get("歧视偏见")),
        "suggest": to_int(r.get("参与建议")),
        "unknownLaw": to_int(r.get("不了解该法律")),
        "other": to_int(pick(r, ["其他/未分类", "其他/专题讨论/未分类"])),
        "supportRate": round(to_float(r.get("支持占比")) * 100, 2),
        "nonSupport": to_int(r.get("非支持意见数")),
    })

total_platform = sum(p["total"] for p in platforms)
total_support = sum(p["support"] for p in platforms)

# 地区统计
region_rows = read_table_rows(summary_file, "地区统计", ["地区分工组/补测省份", "有效公众意见总数"])
region_groups_raw = []
for r in region_rows:
    name = clean_text(pick(r, ["地区分工组/补测省份", "地区分工组"]))
    if not name or name == "合计" or "小计" in name or "合计" in name or "说明" in name:
        continue
    region_groups_raw.append({
        "name": name,
        "total": to_int(r.get("有效公众意见总数")),
        "support": to_int(r.get("支持认可")),
        "neutral": to_int(pick(r, ["中性信息", "中性/观点不明"])),
        "qa": to_int(r.get("咨询疑问")),
        "worry": to_int(r.get("担忧影响")),
        "criticism": to_int(r.get("明确批评")),
        "complaint": to_int(pick(r, ["投诉维权", "投诉维权/举报"])),
        "implement": to_int(r.get("实施问题")),
        "fairness": to_int(r.get("公平争议")),
        "discrimination": to_int(r.get("歧视偏见")),
        "pending": to_int(pick(r, ["待核实", "其他/专题讨论/未分类", "其他/未分类"])),
    })

# 省级态度统计（31个省级行政区，含东部补测）
PROV_FULL = {
    "北京": "北京市", "天津": "天津市", "河北": "河北省", "山西": "山西省", "内蒙古": "内蒙古自治区",
    "辽宁": "辽宁省", "吉林": "吉林省", "黑龙江": "黑龙江省", "上海": "上海市", "江苏": "江苏省",
    "浙江": "浙江省", "安徽": "安徽省", "福建": "福建省", "江西": "江西省", "山东": "山东省",
    "河南": "河南省", "湖北": "湖北省", "湖南": "湖南省", "广东": "广东省", "广西": "广西壮族自治区",
    "海南": "海南省", "重庆": "重庆市", "四川": "四川省", "贵州": "贵州省", "云南": "云南省",
    "西藏": "西藏自治区", "陕西": "陕西省", "甘肃": "甘肃省", "青海": "青海省", "宁夏": "宁夏回族自治区",
    "新疆": "新疆维吾尔自治区", "台湾": "台湾省",
}
prov_rows = read_table_rows(summary_file, "省级态度统计", ["省级行政区", "有效公众意见总数"])
province_data = []
for r in prov_rows:
    name = clean_text(r.get("省级行政区"))
    if not name or name == "地区维度总计" or "合计" in name:
        continue
    full = PROV_FULL.get(name, name + "省")
    province_data.append({
        "name": full,
        "short": name,
        "total": to_int(r.get("有效公众意见总数")),
        "support": to_int(r.get("支持认可")),
        "neutral": to_int(pick(r, ["中性信息", "中性/观点不明"])),
        "qa": to_int(r.get("咨询疑问")),
        "worry": to_int(r.get("担忧影响")),
        "criticism": to_int(r.get("明确批评")),
        "complaint": to_int(pick(r, ["投诉维权", "投诉维权/举报"])),
        "implement": to_int(r.get("实施问题")),
        "fairness": to_int(r.get("公平争议")),
        "discrimination": to_int(r.get("歧视偏见")),
        "suggest": to_int(r.get("参与建议")),
        "other": to_int(pick(r, ["其他/未分类", "其他/专题讨论/未分类"])),
    })

province_values = {p["name"]: p["total"] for p in province_data}

# 省份归属映射（依据“地区统计”分工组/补测省份）
GROUP_PROVINCES = {
    "内蒙古、河北及东北": ["内蒙古自治区", "河北省", "辽宁省", "吉林省", "黑龙江省"],
    "新疆": ["新疆维吾尔自治区"],
    "青海、西藏": ["青海省", "西藏自治区"],
    "四川、甘肃民族地区": ["四川省", "甘肃省"],
    "云南东中部民族地区": ["云南省"],
    "云南西部民族地区": ["云南省"],
    "广东": ["广东省"],
    "广西": ["广西壮族自治区"],
    "海南": ["海南省"],
    "北京": ["北京市"],
    "天津": ["天津市"],
    "河南": ["河南省"],
    "江西": ["江西省"],
    "宁夏、贵州、重庆、湖南、湖北、浙江": ["宁夏回族自治区", "贵州省", "重庆市", "湖南省", "湖北省", "浙江省"],
}

GROUP_KEYWORDS = {
    "内蒙古、河北及东北": ["内蒙古","呼和浩特","包头","乌海","赤峰","通辽","鄂尔多斯","呼伦贝尔","巴彦淖尔","乌兰察布","兴安","锡林郭勒","阿拉善","赛罕","科尔沁","河北","石家庄","唐山","秦皇岛","邯郸","邢台","保定","张家口","承德","沧州","廊坊","衡水","辽宁","吉林","黑龙江","沈阳","大连","长春","哈尔滨","延边","白山","松原","四平","通化","辽源","白城","大庆","齐齐哈尔","牡丹江","佳木斯","绥化","黑河","鸡西","双鸭山","鹤岗","伊春","七台河","大兴安岭","抚顺","本溪","丹东","锦州","营口","阜新","辽阳","盘锦","铁岭","朝阳","葫芦岛"],
    "新疆": ["新疆","乌鲁木齐","伊犁","喀什","和田","阿克苏","巴音郭楞","克孜勒苏","昌吉","博尔塔拉","吐鲁番","哈密","塔城","阿勒泰","克拉玛依","石河子","阿拉尔","图木舒克","五家渠","北屯","铁门关","双河","可克达拉","昆玉","胡杨河","新星"],
    "青海、西藏": ["青海","西宁","海东","海北","海西","黄南","果洛","玉树","西藏","拉萨","日喀则","昌都","林芝","山南","那曲","阿里","雪域高原"],
    "四川、甘肃民族地区": ["四川","成都","阿坝","甘孜","凉山","攀枝花","绵阳","德阳","广元","遂宁","内江","乐山","南充","眉山","宜宾","广安","达州","雅安","巴中","资阳","泸州","自贡","峨边","马边","木里","盐源","冕宁","越西","甘洛","美姑","雷波","金阳","布拖","普格","喜德","昭觉","宁南","会东","会理","德昌","西昌","甘肃","兰州","临夏","甘南","肃北","肃南","东乡","积石山","庆阳","陇南","天水","白银","平凉","定西","酒泉","张掖","嘉峪关","金昌","武威","张家川","天祝","阿克塞"],
    "云南东中部民族地区": ["云南东中部","昆明","曲靖","玉溪","楚雄","红河","文山","元阳","屏边","金平","河口","麻栗坡","马关","富宁","砚山","丘北","广南","峨山","石林","宜良","澄江","通海","华宁","江川","新平","易门"],
    "云南西部民族地区": ["云南西部","保山","临沧","普洱","西双版纳","大理","德宏","怒江","迪庆","丽江","沧源","沧源佤族自治县","兰坪","维西","贡山","福贡","泸水","香格里拉","巍山","漾濞","南涧","弥渡","祥云","宾川","永平","云龙","洱源","剑川","鹤庆","陇川","盈江","梁河","瑞丽","芒市","景洪","勐海","勐腊"],
    "广东、广西、海南": ["广东","广州","深圳","珠海","汕头","佛山","韶关","湛江","肇庆","江门","茂名","惠州","梅州","汕尾","河源","阳江","清远","东莞","中山","潮州","揭阳","云浮","广西","南宁","柳州","桂林","梧州","北海","防城港","钦州","贵港","玉林","百色","贺州","河池","来宾","崇左","海南","海口","三亚","三沙","儋州","五指山","琼海","文昌","万宁","东方","定安","屯昌","澄迈","临高","白沙","昌江","乐东","陵水","保亭","琼中"],
    "宁夏、贵州、重庆、湖南、湖北、浙江": ["宁夏","银川","石嘴山","吴忠","固原","中卫","贵州","贵阳","六盘水","遵义","安顺","毕节","铜仁","黔西南","黔东南","黔南","玉屏","重庆","秀山","石柱","酉阳","彭水","黔江","湖南","长沙","株洲","湘潭","衡阳","邵阳","岳阳","常德","张家界","益阳","郴州","永州","怀化","娄底","湘西","麻阳","湖北","武汉","黄石","十堰","宜昌","襄阳","鄂州","荆门","孝感","荆州","黄冈","咸宁","随州","恩施","浙江","杭州","宁波","温州","嘉兴","湖州","绍兴","金华","衢州","舟山","台州","丽水"],
}

def tag_group(region_text):
    if not region_text:
        return ""
    for group, kws in GROUP_KEYWORDS.items():
        for kw in kws:
            if kw and kw in region_text:
                return group
    return ""

# 拆分广东/广西/海南分组关键词，并补充东部补测省份关键词
GROUP_KEYWORDS["广东"] = ["广东","广州","深圳","珠海","汕头","佛山","韶关","湛江","肇庆","江门","茂名","惠州","梅州","汕尾","河源","阳江","清远","东莞","中山","潮州","揭阳","云浮"]
GROUP_KEYWORDS["广西"] = ["广西","南宁","柳州","桂林","梧州","北海","防城港","钦州","贵港","玉林","百色","贺州","河池","来宾","崇左"]
GROUP_KEYWORDS["海南"] = ["海南","海口","三亚","三沙","儋州","五指山","琼海","文昌","万宁","东方","定安","屯昌","澄迈","临高","白沙","昌江","乐东","陵水","保亭","琼中"]
GROUP_KEYWORDS["北京"] = ["北京","东城","西城","朝阳区","海淀","丰台","石景山","通州","昌平","大兴","顺义","房山","门头沟","怀柔","密云","延庆","平谷"]
GROUP_KEYWORDS["天津"] = ["天津","和平区","河东","河西","南开","河北区","红桥","滨海","东丽","西青","津南","北辰","武清","宝坻","静海","宁河","蓟州"]
GROUP_KEYWORDS["河南"] = ["河南","郑州","开封","洛阳","平顶山","安阳","鹤壁","新乡","焦作","濮阳","许昌","漯河","三门峡","南阳","商丘","信阳","周口","驻马店"]
GROUP_KEYWORDS["江西"] = ["江西","南昌","景德镇","萍乡","九江","新余","鹰潭","赣州","吉安","宜春","抚州","上饶"]
GROUP_KEYWORDS.pop("广东、广西、海南", None)

province_groups = {}
for g in region_groups_raw:
    name = g["name"]
    total = g["total"]
    for prov in GROUP_PROVINCES.get(name, []):
        province_groups.setdefault(prov, []).append({"group": name, "total": total})

# 监测来源 / 查看信息（新表：层级 | 平台组/地区 | 数量 | 说明，取“原平台组小计”）
def find_sheet_total(sheet, lvl="原平台组", nm="小计"):
    rows = load_sheet(summary_file, sheet)
    for r in rows:
        joined = "|".join(str(c) for c in r if c is not None)
        if lvl in joined and nm in joined:
            for c in r:
                if isinstance(c, (int, float)) and c > 0:
                    return int(c)
            for c in r:
                m = re.search(r"\d+", str(c))
                if m:
                    return int(m.group())
    return 0

total_sources = find_sheet_total("监测来源统计")
total_views = find_sheet_total("查看信息统计")

# 非支持态度统计（平台维度 305）
non_support_rows = read_table_rows(summary_file, "非支持态度统计", ["类别", "数量"])
non_support = []
for r in non_support_rows:
    name = clean_text(r.get("类别"))
    if not name or "合计" in name or "小计" in name:
        continue
    name = name.replace("投诉维权/举报", "投诉维权").replace("中性/观点不明", "中性信息")
    non_support.append({"name": name, "value": to_int(r.get("数量"))})
non_support_total = sum(n["value"] for n in non_support)

# 语言统计（位置式解析：平台维度=第1/2列，地区维度=第4/5列）
lang_rows = load_sheet(summary_file, "语言统计")
lang_plain = [r for r in lang_rows if any(c is not None and str(c).strip() != "" for c in r)]
language_platform = []
language_region = []
for r in lang_plain:
    if len(r) < 2 or not isinstance(r[0], str):
        continue
    plat_name = r[0].strip()
    if "合计" in plat_name:
        continue
    plat_val = r[1] if len(r) > 1 and isinstance(r[1], (int, float)) else None
    if plat_val is not None:
        language_platform.append({"name": plat_name, "value": int(plat_val)})
    if len(r) >= 5 and isinstance(r[3], str) and isinstance(r[4], (int, float)):
        reg_name = r[3].strip()
        language_region.append({"name": reg_name, "value": int(r[4])})

minority_lang_count = sum(x["value"] for x in language_platform if not x["name"].startswith("中文") and "合计" not in x["name"])

# ---------------- 2. 明细：评论原话 / 时间趋势 / 高热内容 ----------------
quotes = []
trend = {}
hot_items = []
quote_keys = set()

def add_quote(platform, region, attitude, text, date, source, account="", language="中文"):
    t = clean_text(text)
    if not is_meaningful_quote(t):
        return
    if attitude in ("无关评论", "已排除", "不计入", "排除记录"):
        return
    if "官方" in attitude or "媒体" in attitude or "报道" in attitude or "宣传" in attitude or "基层执行" in attitude or "依法履职" in attitude:
        return
    att_map = {
        "质疑": "明确批评", "批评": "明确批评", "担忧": "担忧影响", "咨询": "咨询疑问",
        "咨询疑问/诉求": "咨询疑问", "中性/观点不明": "中性信息", "投诉维权/举报": "投诉维权",
        "其他/专题讨论/未分类": "其他/未分类", "咨询/诉求": "咨询疑问",
        "待维吾尔语人员复核": "待核实", "支持认可促进法": "支持认可",
    }
    attitude = att_map.get(attitude, attitude)
    if t in quote_keys:
        return
    quote_keys.add(t)
    if region == "未展示":
        region = ""
    quotes.append({
        "platform": platform,
        "region": region,
        "group": tag_group(region),
        "attitude": attitude,
        "text": t,
        "date": date.isoformat() if isinstance(date, datetime.date) else "",
        "source": source,
        "src": source,
        "account": account,
        "language": language,
    })

def count_trend(date):
    if isinstance(date, datetime.date):
        key = date.isoformat()
        trend[key] = trend.get(key, 0) + 1

def is_public_attitude(att):
    if not att:
        return False
    if any(k in att for k in ("官方", "媒体", "报道", "宣传", "会议", "普法", "基层执行", "依法履职")):
        return False
    return True

def scan_detail(path, sheets, date_cols, count_cols, quote_cols, plat_cols, region_cols, att_cols, account_cols=None, lang_cols=None):
    for sheet in sheets:
        try:
            rows = load_sheet(path, sheet)
        except Exception:
            continue
        non_empty = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
        if not non_empty:
            continue
        # 找表头
        hidx = None
        for i, r in enumerate(non_empty[:8]):
            joined = "|".join(str(c) for c in r if c is not None)
            if ("发布日期" in joined or "评论日期" in joined or "序号" in joined) and ("平台" in joined or "意见" in joined or "来源" in joined):
                hidx = i
                break
        if hidx is None:
            continue
        header = non_empty[hidx]
        def ci(names):
            return col_index(header, names)
        i_date = ci(date_cols)
        i_count = ci(count_cols)
        i_quote = ci(quote_cols)
        i_plat = ci(plat_cols)
        i_region = ci(region_cols)
        i_att = ci(att_cols)
        i_acc = ci(account_cols) if account_cols else -1
        i_lang = ci(lang_cols) if lang_cols else -1
        if i_date < 0 and i_quote < 0:
            continue
        for r in non_empty[hidx+1:]:
            counted = False
            if i_count >= 0 and i_count < len(r):
                cv = r[i_count]
                counted = cv is not None and str(cv).strip() in ("是", "1", "TRUE", "True", "计入", "已计入")
            else:
                # 未提供计入列时，按明细表默认计入（用于公众意见明细表）
                counted = True
            date = parse_date(r[i_date]) if i_date >= 0 and i_date < len(r) else None
            if counted and date:
                count_trend(date)
            quote = clean_text(r[i_quote]) if i_quote >= 0 and i_quote < len(r) else ""
            if not quote:
                continue
            plat = clean_text(r[i_plat]) if i_plat >= 0 and i_plat < len(r) else ""
            reg = clean_text(r[i_region]) if i_region >= 0 and i_region < len(r) else ""
            att = clean_text(r[i_att]) if i_att >= 0 and i_att < len(r) else ""
            acc = clean_text(r[i_acc]) if i_acc >= 0 and i_acc < len(r) else ""
            lang = clean_text(r[i_lang]) if i_lang >= 0 and i_lang < len(r) else "中文"
            if not is_public_attitude(att):
                continue
            if not plat and not att:
                continue
            add_quote(plat or "其他", reg, att, quote, date, os.path.basename(path), acc, lang)

# 非支持来源与原话（新文件：216 条证据全量，含地区与日期）
nonsup_rows = load_sheet(summary_file, "非支持来源与原话")
nonsup_plain = [r for r in nonsup_rows if any(c is not None and str(c).strip() != "" for c in r)]
hidx = None
for i, r in enumerate(nonsup_plain[:12]):
    joined = "|".join(str(c) for c in r if c is not None)
    if "原话" in joined and "平台" in joined and "序号" in joined:
        hidx = i
        break
if hidx is not None:
    header = nonsup_plain[hidx]
    i_reg = col_index(header, ["来源表地区/涉及地区"])
    i_plat = col_index(header, ["平台"])
    i_acc = col_index(header, ["账号/栏目"])
    i_att = col_index(header, ["态度类别"])
    i_q = col_index(header, ["原话/代表性原话"])
    i_date = col_index(header, ["发布时间/采集日期"])
    i_status = col_index(header, ["原表计入状态"])
    for r in nonsup_plain[hidx+1:]:
        plat = clean_text(r[i_plat]) if 0 <= i_plat < len(r) else ""
        reg = clean_text(r[i_reg]) if 0 <= i_reg < len(r) else ""
        acc = clean_text(r[i_acc]) if 0 <= i_acc < len(r) else ""
        att = clean_text(r[i_att]) if 0 <= i_att < len(r) else ""
        att = att.replace("质疑", "明确批评").replace("投诉维权/举报", "投诉维权").replace("中性/观点不明", "中性信息")
        q = clean_text(r[i_q]) if 0 <= i_q < len(r) else ""
        date = parse_date(r[i_date]) if 0 <= i_date < len(r) else None
        status = clean_text(r[i_status]) if 0 <= i_status < len(r) else ""
        counted = "计入" in status or "有效" in status
        if not q or not att:
            continue
        if plat in ("多个账号/评论区", "合计"):
            continue
        if counted and date:
            count_trend(date)
        add_quote(plat or "其他", reg, att, q, date, "非支持来源与原话(全量)", acc)

# 四川甘肃 公众评论明细
scan_detail(
    os.path.join(ROOT, "中华人民共和国民族团结进步促进法_四川甘肃民族地区舆情监测汇总_复核完成.xlsx"),
    ["公众评论明细"],
    ["评论日期", "母帖发布日期"],
    ["是否计入公众统计"],
    ["评论原文"],
    ["平台"],
    ["涉及地区", "评论者公开地区"],
    ["意见类型"],
    account_cols=["评论者公开昵称"],
    lang_cols=["原始语言"],
)

# 快手
scan_detail(
    os.path.join(ROOT, "快手平台民族团结进步促进法舆情监测_统计修正版.xlsx"),
    ["有效舆情明细", "扁平化数据"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
    lang_cols=["原始语言"],
)

# 抖音
scan_detail(
    os.path.join(ROOT, "抖音等视频平台_民族团结进步促进法舆情监测_最终修正版.xlsx"),
    ["有效舆情明细"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
    lang_cols=["原始语言"],
)

# 微博热榜
scan_detail(
    os.path.join(ROOT, "微博热榜平台_民族团结进步促进法舆情监测_最终修正版.xlsx"),
    ["有效舆情明细表", "数据明细汇总"],
    ["发布日期"],
    ["是否重复"],
    ["评论观点", "微博原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号名称/发布单位", "账号或栏目名称"],
)

# Reddit/微信15地区
scan_detail(
    os.path.join(ROOT, "15地区微信及Reddit舆情监测_最终修正版.xlsx"),
    ["公众舆情明细"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["中文译文", "原文证据摘录"],
    ["平台/网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
    lang_cols=["原始语言"],
)

# 新疆区州
scan_detail(
    os.path.join(ROOT, "新疆区州网页与视频号舆情监测_最终修正版.xlsx"),
    ["第一部分_有效舆情明细"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
    lang_cols=["原始语言"],
)

# 知乎/B站/百度知道
scan_detail(
    os.path.join(ROOT, "知乎_哔哩哔哩_百度知道舆情监测_最终修正版.xlsx"),
    ["扁平化数据", "有效舆情明细"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录", "客观摘要"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
)

# 内蒙古河北东北
scan_detail(
    os.path.join(ROOT, "内蒙古河北东北民族地区_舆情监测统计修正版(2).xlsx"),
    ["有效舆情明细表", "扁平化数据表"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["统计地区", "涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
    lang_cols=["原始语言"],
)

# 云南东中部
scan_detail(
    os.path.join(ROOT, "云南东中部民族地区_舆情监测_最终修正版.xlsx"),
    ["有效舆情明细表", "扁平化数据"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
)

# 第二组16县
scan_detail(
    os.path.join(ROOT, "第二组16县_民族团结进步促进法舆情回溯_最终修正版.xlsx"),
    ["有效舆情明细"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
)

# 青海西藏
scan_detail(
    os.path.join(ROOT, "青海西藏舆情省级拆分_修正版.xlsx"),
    ["扁平化数据（青海）", "扁平化数据（西藏）"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["统计地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
    lang_cols=["原始语言"],
)

# 新疆18平台
scan_detail(
    os.path.join(ROOT, "新疆民族团结进步促进法_18平台舆情监测_最终去重修正版.xlsx"),
    ["明细"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
)

# 新疆地区及全国性平台补充
scan_detail(
    os.path.join(ROOT, "新疆地区及全国性平台补充_民族团结进步促进法舆情监测_最终修正版.xlsx"),
    ["有效舆情明细"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
)

# 多语言专项复核
scan_detail(
    os.path.join(ROOT, "民族团结进步促进法_多语言专项复核_最终修正版.xlsx"),
    ["有效舆情明细"],
    ["发布日期"],
    ["是否计入公众意见统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
    lang_cols=["原始语言"],
)

# 舆情监测报告统计口径修正版
scan_detail(
    os.path.join(ROOT, "民族团结进步促进法舆情监测报告_统计口径修正版.xlsx"),
    ["有效舆情明细", "扁平化数据"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
)

# 云南地区（采集记录模板）
scan_detail(
    os.path.join(ROOT, "云南地区民族团结进步促进法舆情监测_最终修正版.xlsx"),
    ["舆情采集记录模板"],
    ["发布日期"],
    ["是否计入公众统计"],
    ["原文证据摘录"],
    ["平台或网站", "平台"],
    ["涉及地区"],
    ["意见类型"],
    account_cols=["账号或栏目名称"],
)

# 高热内容榜：从明细表提取有真实互动量的内容
def parse_interact(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", "").strip()
    s = re.sub(r"[（(].*?[)）]", "", s)
    m = re.search(r"([\d.]+)\s*(万|w|W)?", s)
    if not m:
        return None
    val = float(m.group(1))
    if m.group(2):
        val *= 10000
    return int(val)

def scan_hot(path, sheets):
    for sheet in sheets:
        try:
            rows = load_sheet(path, sheet)
        except Exception:
            continue
        non_empty = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
        if not non_empty:
            continue
        hidx = None
        for i, r in enumerate(non_empty[:8]):
            joined = "|".join(str(c) for c in r if c is not None)
            if "点赞" in joined and "平台" in joined:
                hidx = i
                break
        if hidx is None:
            continue
        header = non_empty[hidx]
        i_like = col_index(header, ["点赞量"])
        i_cmt = col_index(header, ["评论量"])
        i_rt = col_index(header, ["转发量"])
        i_plat = col_index(header, ["平台或网站", "平台"])
        i_acc = col_index(header, ["账号或栏目名称", "账号名称/发布单位", "母帖账号"])
        i_title = col_index(header, ["标题或话题", "母帖标题"])
        i_date = col_index(header, ["发布日期", "评论日期"])
        for r in non_empty[hidx+1:]:
            likes = parse_interact(r[i_like]) if 0 <= i_like < len(r) else None
            if not likes:
                continue
            hot_items.append({
                "platform": clean_text(r[i_plat]) if 0 <= i_plat < len(r) else "",
                "account": clean_text(r[i_acc]) if 0 <= i_acc < len(r) else "",
                "title": clean_text(r[i_title]) if 0 <= i_title < len(r) else "",
                "likes": likes,
                "comments": parse_interact(r[i_cmt]) if 0 <= i_cmt < len(r) else None,
                "shares": parse_interact(r[i_rt]) if 0 <= i_rt < len(r) else None,
                "date": (parse_date(r[i_date]).isoformat() if i_date >= 0 and i_date < len(r) and parse_date(r[i_date]) else ""),
            })

for f, sheets in [
    (os.path.join(ROOT, "快手平台民族团结进步促进法舆情监测_统计修正版.xlsx"), ["有效舆情明细"]),
    (os.path.join(ROOT, "抖音等视频平台_民族团结进步促进法舆情监测_最终修正版.xlsx"), ["有效舆情明细"]),
    (os.path.join(ROOT, "微博热榜平台_民族团结进步促进法舆情监测_最终修正版.xlsx"), ["有效舆情明细表"]),
    (os.path.join(ROOT, "中华人民共和国民族团结进步促进法_四川甘肃民族地区舆情监测汇总_复核完成.xlsx"), ["公众评论明细"]),
    (os.path.join(ROOT, "15地区微信及Reddit舆情监测_最终修正版.xlsx"), ["公众舆情明细"]),
]:
    scan_hot(f, sheets)

hot_items.sort(key=lambda x: x["likes"], reverse=True)
deduped = []
seen_hot = set()
for h in hot_items:
    key = (h["account"], h["title"], h["likes"])
    if key in seen_hot:
        continue
    seen_hot.add(key)
    deduped.append(h)
hot_top = deduped[:9]

# 原话池按信息量排序，让滚动流以更充实的内容开场
quotes.sort(key=lambda q: (len(q["text"]), q["attitude"] == "支持认可"), reverse=True)

# 时间趋势按日期排序
trend_series = [{"date": d, "value": v} for d, v in sorted(trend.items())]

# 语言轮播代表记录（来自少数民族语言复核等真实记录）
lang_spotlight = [
    {"lang": "维吾尔语/文", "record": "快手「麦谷」发布维吾尔语教学视频，响应民族团结进步促进法（支持认可）", "count": 1},
    {"lang": "彝语/文", "record": "快手「放猪老人~迪日尔古」彝语口播普法视频，讲解禁止招聘民族歧视条款", "count": 1},
    {"lang": "藏语/文", "record": "视频号「甘南融媒藏语平台」藏语普法母帖，4条评论均为表情型支持", "count": 0},
    {"lang": "中文/普通话", "record": "全网监测主体为中文评论；含大量支持认可与咨询、担忧等声音", "count": 12349},
]

province_list = []
for p in province_data:
    prov = {"name": p["name"], "short": p["short"], "value": p["total"],
            "groups": province_groups.get(p["name"], [])}
    for k in ("support","neutral","qa","worry","criticism","complaint","implement","fairness","discrimination","suggest","other"):
        prov[k] = p[k]
    province_list.append(prov)
province_list.sort(key=lambda x: x["value"], reverse=True)

# 地区轮播维度：省级（每个省份一条，地图高亮单个省）
province_regions = []
for p in province_list:
    groups = p.get("groups") or []
    province_regions.append({
        "name": p["short"],
        "total": p["value"],
        "support": p["support"],
        "neutral": p["neutral"],
        "qa": p["qa"],
        "worry": p["worry"],
        "criticism": p["criticism"],
        "complaint": p["complaint"],
        "implement": p["implement"],
        "fairness": p["fairness"],
        "discrimination": p["discrimination"],
        "pending": 0,
        "provinces": [p["short"]],
        "sourceGroup": groups[0]["group"] if groups else "",
    })

data = {
    "generatedAt": "2026-08-11",
    "topStats": {
        "totalOpinions": total_platform,
        "totalOpinionsLabel": "平台维度公众意见总量",
        "regionOpinions": sum(g["total"] for g in region_groups_raw),
        "regionOpinionsLabel": "可明确归属地区的公众意见",
        "supportCount": total_support,
        "supportRate": round(total_support / total_platform * 100, 2),
        "regionSupport": sum(g["support"] for g in region_groups_raw),
        "regionSupportRate": round(sum(g["support"] for g in region_groups_raw) / sum(g["total"] for g in region_groups_raw) * 100, 2) if sum(g["total"] for g in region_groups_raw) else 0,
        "nonSupport": non_support_total,
        "nonSupportLabel": "非支持/非肯定态度",
        "monitorSources": total_sources,
        "monitorSourcesLabel": "监测来源数",
        "viewedInfo": total_views,
        "viewedInfoLabel": "查看信息数",
        "minorityLang": minority_lang_count,
        "minorityLangLabel": "少数民族语言舆情",
    },
    "platforms": platforms,
    "languagePlatform": language_platform,
    "languageRegion": language_region,
    "langSpotlight": lang_spotlight,
    "regions": province_regions,
    "provinces": province_list,
    "attitude": {
        "macro": [
            {"name": "支持认可", "value": total_support},
            {"name": "中性信息", "value": sum(p["neutral"] for p in platforms)},
            {"name": "参与建议", "value": sum(p["suggest"] for p in platforms)},
            {"name": "非支持/非肯定", "value": non_support_total},
        ],
        "detail": [
            {"name": "支持认可", "value": total_support},
            {"name": "中性信息", "value": sum(p["neutral"] for p in platforms)},
            {"name": "咨询疑问", "value": sum(p["qa"] for p in platforms)},
            {"name": "担忧影响", "value": sum(p["worry"] for p in platforms)},
            {"name": "明确批评", "value": sum(p["criticism"] for p in platforms)},
            {"name": "投诉维权", "value": sum(p["complaint"] for p in platforms)},
            {"name": "实施问题", "value": sum(p["implement"] for p in platforms)},
            {"name": "公平争议", "value": sum(p["fairness"] for p in platforms)},
            {"name": "歧视偏见", "value": sum(p["discrimination"] for p in platforms)},
            {"name": "参与建议", "value": sum(p["suggest"] for p in platforms)},
            {"name": "不了解该法律", "value": sum(p["unknownLaw"] for p in platforms)},
        ],
    },
    "nonSupport": non_support,
    "trend": trend_series,
    "quotes": quotes,
    "hotTop": hot_top,
}

write_outputs(data)

print("平台:", len(platforms), "合计:", total_platform)
print("地区组:", len(region_groups_raw), "合计:", sum(g["total"] for g in region_groups_raw), "支持:", sum(g["support"] for g in region_groups_raw))
print("省级行政区:", len(province_data), "合计:", sum(p["total"] for p in province_data))
print("监测来源:", total_sources, "查看信息:", total_views, "非支持:", non_support_total, "少语:", minority_lang_count)
print("语言(平台):", language_platform)
print("评论原话池:", len(quotes))
print("趋势天数:", len(trend_series), "总记录:", sum(t["value"] for t in trend_series))
print("高热内容:", len(hot_items), "取前", len(hot_top))
