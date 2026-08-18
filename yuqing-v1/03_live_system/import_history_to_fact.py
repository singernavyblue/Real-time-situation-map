#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""历史 Excel 全量导入舆情事实表

读取 yuqing-v1/02_data/ 下各文件的“有效舆情明细 / 扁平化数据 / 公众舆情明细 /
公众评论明细”等明细表，统一映射成舆情事实表 24 字段，去重后写入原子化工作簿。

用法：
  python import_history_to_fact.py --xlsx 路径/Excel数据库改造示例.xlsx
  python import_history_to_fact.py --xlsx 路径/库.xlsx --dry-run
  python import_history_to_fact.py --xlsx 路径/库.xlsx --no-clear

默认会清空舆情事实表（含待清洗区）的旧数据后再导入；--no-clear 则不清空。
"""
import argparse
import hashlib
import json
import os
import re
import shutil
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from clean_and_append import _parse_dt, _to_int, find_header_row, column_index, FACT_FIELDS
from ingest import is_minority_lang, platform_group

PROVINCES = [
    "北京", "天津", "河北", "山西", "内蒙古", "辽宁", "吉林", "黑龙江", "上海",
    "江苏", "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南",
    "广东", "广西", "海南", "重庆", "四川", "贵州", "云南", "西藏", "陕西",
    "甘肃", "青海", "宁夏", "新疆", "台湾",
]
PROV_FULL = {
    "北京市": "北京", "天津市": "天津", "河北省": "河北", "山西省": "山西",
    "内蒙古自治区": "内蒙古", "辽宁省": "辽宁", "吉林省": "吉林", "黑龙江省": "黑龙江",
    "上海市": "上海", "江苏省": "江苏", "浙江省": "浙江", "安徽省": "安徽",
    "福建省": "福建", "江西省": "江西", "山东省": "山东", "河南省": "河南",
    "湖北省": "湖北", "湖南省": "湖南", "广东省": "广东", "广西壮族自治区": "广西",
    "海南省": "海南", "重庆市": "重庆", "四川省": "四川", "贵州省": "贵州",
    "云南省": "云南", "西藏自治区": "西藏", "陕西省": "陕西", "甘肃省": "甘肃",
    "青海省": "青海", "宁夏回族自治区": "宁夏", "新疆维吾尔自治区": "新疆",
    "台湾省": "台湾",
}

ETHNIC_WORDS = (
    "维吾尔", "哈萨克", "柯尔克孜", "塔吉克", "锡伯", "达斡尔", "鄂温克", "鄂伦春",
    "蒙古", "回", "藏", "苗", "壮", "彝", "瑶", "侗", "满", "土家", "畲", "黎",
    "水族", "白族", "傣", "景颇", "拉祜", "佤", "纳西", "独龙", "怒", "傈僳",
    "普米", "布朗", "阿昌", "德昂", "基诺", "保安", "东乡", "撒拉", "裕固",
    "羌", "朝鲜", "土族", "仫佬", "毛南",
)

# 常见市县/地区 -> 省份（补 site_sources 之外的城市级关键词）
CITY_PROVINCE = {
    # 新疆
    "乌鲁木齐": "新疆", "哈密": "新疆", "伊犁": "新疆", "伊宁": "新疆", "奎屯": "新疆",
    "昌吉": "新疆", "博州": "新疆", "博尔塔拉": "新疆", "巴州": "新疆", "巴音郭楞": "新疆",
    "克州": "新疆", "克孜勒苏": "新疆", "阿克苏": "新疆", "和田": "新疆", "喀什": "新疆",
    "塔城": "新疆", "阿勒泰": "新疆", "吐鲁番": "新疆", "克拉玛依": "新疆", "石河子": "新疆",
    "五家渠": "新疆", "阿拉尔": "新疆", "图木舒克": "新疆", "木垒": "新疆", "巴里坤": "新疆",
    "察布查尔": "新疆", "焉耆": "新疆", "和布克赛尔": "新疆", "塔什库尔干": "新疆",
    "乌什": "新疆", "英吉沙": "新疆", "疏附": "新疆", "霍城": "新疆", "和硕": "新疆",
    "双河": "新疆", "第五师": "新疆", "兵团": "新疆",
    # 云南
    "大理": "云南", "楚雄": "云南", "红河": "云南", "文山": "云南", "西双版纳": "云南",
    "景洪": "云南", "勐腊": "云南", "德宏": "云南", "怒江": "云南", "迪庆": "云南",
    "丽江": "云南", "保山": "云南", "临沧": "云南", "普洱": "云南", "西盟": "云南",
    "曲靖": "云南", "玉溪": "云南", "昭通": "云南", "昆明": "云南", "祥云": "云南",
    "永平": "云南", "洱源": "云南", "漾濞": "云南", "峨山": "云南", "麻栗坡": "云南",
    "元阳": "云南", "屏边": "云南", "金平": "云南",
    # 内蒙古
    "呼伦贝尔": "内蒙古", "通辽": "内蒙古", "锡林郭勒": "内蒙古", "科尔沁": "内蒙古",
    "赤峰": "内蒙古", "乌兰察布": "内蒙古", "巴彦淖尔": "内蒙古", "鄂尔多斯": "内蒙古",
    "包头": "内蒙古", "乌海": "内蒙古", "兴安": "内蒙古", "阿拉善": "内蒙古", "呼和浩特": "内蒙古",
    # 其他省
    "阜新": "辽宁", "喀喇沁左翼": "辽宁",
    "杜尔伯特": "黑龙江",
    "宝鸡": "陕西",
    "甘南": "甘肃", "临夏": "甘肃", "积石山": "甘肃", "东乡": "甘肃", "天祝": "甘肃",
    "张家川": "甘肃", "阿克塞": "甘肃", "肃南": "甘肃", "肃北": "甘肃",
    "上饶": "江西", "南昌": "江西",
    "珠三角": "广东",
    "延边": "吉林", "长白": "吉林", "前郭": "吉林", "伊通": "吉林",
    "恩施": "湖北", "长阳": "湖北", "五峰": "湖北",
    "湘西": "湖南", "城步": "湖南", "江华": "湖南", "通道": "湖南", "新晃": "湖南",
    "芷江": "湖南", "靖州": "湖南", "麻阳": "湖南",
    "黔东南": "贵州", "黔南": "贵州", "黔西南": "贵州", "玉屏": "贵州", "威宁": "贵州",
    "松桃": "贵州", "印江": "贵州", "道真": "贵州", "务川": "贵州", "沿河": "贵州",
    "镇宁": "贵州", "紫云": "贵州", "关岭": "贵州", "三都": "贵州",
    "阿坝": "四川", "凉山": "四川", "甘孜": "四川", "木里": "四川", "马边": "四川",
    "峨边": "四川", "北川": "四川",
    "海南藏族自治州": "青海", "海南州": "青海", "海北": "青海", "海西": "青海",
    "黄南": "青海", "果洛": "青海", "玉树": "青海", "门源": "青海", "化隆": "青海",
    "互助": "青海", "循化": "青海", "河南蒙古族自治县": "青海", "大通": "青海", "民和": "青海",
    "石柱": "重庆", "秀山": "重庆", "酉阳": "重庆", "彭水": "重庆",
    "青龙": "河北", "大厂": "河北", "孟村": "河北", "丰宁": "河北", "围场": "河北", "宽城": "河北",
    "连南": "广东", "连山": "广东", "乳源": "广东",
    "白沙": "海南", "昌江": "海南", "乐东": "海南", "陵水": "海南", "保亭": "海南", "琼中": "海南",
    "景宁": "浙江",
}


def _cell(v):
    return "" if v is None else str(v).strip()


def _clean_text(v):
    return re.sub(r"\s+", " ", _cell(v)).strip()


def _gen_uid(platform, published, account, text):
    key = "|".join([platform, published, account, text])
    return "hist-" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


def _load_province_keywords(site_sources_path):
    """从 site_sources.json 提取 自治州/县名 -> 省份 关键词表。"""
    kw = {}
    try:
        with open(site_sources_path, encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = {"sources": []}
    for s in data.get("sources", []):
        name = _clean_text(s.get("name"))
        prov = _clean_text(s.get("province"))
        if not name or not prov:
            continue
        kw[name] = prov
        stem = re.sub(r"(自治州|自治县|自治旗|地区|市|县|旗|区|省)$", "", name)
        if len(stem) >= 2:
            kw.setdefault(stem, prov)
        for w in ETHNIC_WORDS:
            stem2 = stem.replace(w, "")
            if len(stem2) >= 2:
                kw.setdefault(stem2, prov)
    for k, v in CITY_PROVINCE.items():
        kw.setdefault(k, v)
    return kw


def _infer_province(region, region_group, province_col, kw_map):
    if province_col:
        for short, full in PROV_FULL.items():
            if short in province_col or province_col in short:
                return PROV_FULL.get(province_col, province_col)
        for p in PROVINCES:
            if p in province_col:
                return p
    for text in (region_group, region):
        if not text:
            continue
        for name, prov in sorted(kw_map.items(), key=lambda kv: -len(kv[0])):
            if name and name in text:
                return prov
        for full, short in PROV_FULL.items():
            if full in text:
                return short
        for p in PROVINCES:
            if p in text:
                return p
    return ""


def find_detail_header(rows):
    """普通明细表：发布日期 + 原文/标题 + 平台。"""
    for r in rows[:12]:
        vals = [_cell(c) for c in r]
        joined = "|".join(vals)
        if (
            sum(1 for x in vals if x) >= 3
            and "发布日期" in joined
            and ("原文证据摘录" in joined or "标题" in joined)
            and "平台" in joined
            and "排除原因" not in joined
        ):
            return r
    return None


def find_comment_header(rows):
    """公众评论明细：评论原文 + 评论日期/母帖 + 平台。"""
    for r in rows[:12]:
        vals = [_cell(c) for c in r]
        joined = "|".join(vals)
        if (
            sum(1 for x in vals if x) >= 3
            and "评论原文" in joined
            and ("评论日期" in joined or "母帖" in joined)
            and "平台" in joined
        ):
            return r
    return None


def find_template_header(rows):
    """舆情采集记录模板：公众原始意见摘录 + 平台/网站。"""
    for r in rows[:12]:
        vals = [_cell(c) for c in r]
        joined = "|".join(vals)
        if (
            sum(1 for x in vals if x) >= 3
            and "公众原始意见摘录" in joined
            and "平台/网站" in joined
        ):
            return r
    return None


def find_nonsupport_header(rows):
    """非支持来源与原话：原话/代表性原话 + 平台 + 态度类别。"""
    for r in rows[:16]:
        vals = [_cell(c) for c in r]
        joined = "|".join(vals)
        if (
            sum(1 for x in vals if x) >= 3
            and ("原话/代表性原话" in joined or "原话" in joined)
            and "平台" in joined
            and "态度类别" in joined
        ):
            return r
    return None


def col_index_map(header):
    m = {}
    for i, h in enumerate(header):
        t = _cell(h)
        if not t:
            continue
        if "发布日期" in t or "发布时间" in t:
            m.setdefault("发布时间", i)
        elif "采集日期" in t:
            m.setdefault("采集时间", i)
        elif "平台/网站" in t or "平台或网站" in t or t == "平台":
            m.setdefault("平台", i)
        elif "采集来源" in t:
            m.setdefault("具体来源", i)
        elif "账号或栏目" in t or "账号或发布" in t or "账号/栏目" in t or t == "账号":
            m.setdefault("账号", i)
        elif "原文证据摘录" in t or "公众原始意见摘录" in t or "原话/代表性原话" in t:
            m.setdefault("正文", i)
        elif "标题/主题" in t or "标题或话题" in t or "标题或主题" in t or "文件或活动名称" in t or "母帖标题" in t or "母帖/母视频/标题" in t:
            m.setdefault("标题", i)
        elif "客观摘要" in t or "内容摘要" in t:
            m.setdefault("摘要", i)
        elif "中文译文" in t or "中文翻译" in t:
            m.setdefault("译文", i)
        elif "原帖/原文链接" in t or "原始链接" in t or "原文链接" in t or "母帖原始链接" in t or "原始链接/证据入口" in t:
            m.setdefault("原始链接", i)
        elif "统计地区" in t or "地区大类" in t:
            m.setdefault("地区组", i)
        elif "省份" in t:
            m.setdefault("省份", i)
        elif "涉及地区" in t:
            m.setdefault("地区", i)
        elif "评论者公开地区" in t:
            m.setdefault("评论者地区", i)
        elif "原始语言" in t:
            m.setdefault("语言", i)
        elif "意见类型" in t or "态度类别" in t:
            m.setdefault("态度", i)
        elif "具体议题" in t or "问题议题" in t:
            m.setdefault("议题", i)
        elif "点赞" in t:
            m.setdefault("点赞", i)
        elif "评论" in t and "评论总数" not in t:
            m.setdefault("评论", i)
        elif "转发" in t:
            m.setdefault("转发", i)
        elif "来源类型" in t:
            m.setdefault("来源类型", i)
        elif "关联度" in t:
            m.setdefault("关联度", i)
        elif "是否重复" in t:
            m.setdefault("是否重复", i)
        elif "是否待核实" in t or "是否需要复核" in t:
            m.setdefault("是否待核实", i)
        elif "是否计入公众统计" in t:
            m.setdefault("是否计入", i)
        elif "原表计入状态" in t:
            m.setdefault("是否计入", i)
        elif "评论日期" in t:
            m.setdefault("评论日期", i)
        elif "母帖发布日期" in t:
            m.setdefault("母帖日期", i)
        elif "母帖账号" in t:
            m.setdefault("母帖账号", i)
        elif "评论者公开昵称" in t:
            m.setdefault("评论昵称", i)
        elif "评论原文" in t:
            m.setdefault("评论正文", i)
        elif "评论点赞量" in t:
            m.setdefault("点赞", i)
        elif "信息量" in t:
            m.setdefault("信息量", i)
        elif "复核说明" in t:
            m.setdefault("复核说明", i)
        elif "证据来源" in t:
            m.setdefault("证据来源", i)
        elif "检索关键词" in t:
            m.setdefault("检索关键词", i)
        elif "数据编号" in t:
            m.setdefault("数据编号", i)
        elif "证据状态" in t:
            m.setdefault("证据状态", i)
        elif "记录类型" in t:
            m.setdefault("记录类型", i)
        elif "重复来源数" in t:
            m.setdefault("重复来源数", i)
        elif "原始来源文件" in t:
            m.setdefault("原始来源文件", i)
        elif "原始子表" in t:
            m.setdefault("原始子表", i)
    return m


def pick(row, m, key):
    i = m.get(key)
    return _clean_text(row[i]) if i is not None and i < len(row) else ""


def normalize_detail(row, m, fname, kw_map, collected_default):
    platform = pick(row, m, "平台")
    text = pick(row, m, "正文") or pick(row, m, "标题") or pick(row, m, "摘要") or pick(row, m, "译文")
    if not platform or not text:
        return None
    published = pick(row, m, "发布时间") or pick(row, m, "评论日期") or pick(row, m, "母帖日期")
    published = _parse_dt(published)
    account = pick(row, m, "账号") or pick(row, m, "评论昵称") or pick(row, m, "母帖账号")
    region = pick(row, m, "地区") or pick(row, m, "评论者地区")
    region_group = pick(row, m, "地区组")
    province = pick(row, m, "省份")
    if not province:
        province = _infer_province(region, region_group, "", kw_map)
    attitude = pick(row, m, "态度") or "待核实"
    language = pick(row, m, "语言") or "中文"
    counted = pick(row, m, "是否计入")
    if counted in ("否", "0", "不计入", "不纳入"):
        counted_flag = "否"
    else:
        counted_flag = "是"
    notes_parts = [
        f"历史导入:{os.path.basename(fname)}",
        f"来源类型:{pick(row, m, '来源类型')}" if pick(row, m, "来源类型") else "",
        f"信息量:{pick(row, m, '信息量')}" if pick(row, m, "信息量") else "",
        f"复核:{pick(row, m, '复核说明')}" if pick(row, m, "复核说明") else "",
        f"证据:{pick(row, m, '证据来源')}" if pick(row, m, "证据来源") else "",
        f"重复:{pick(row, m, '是否重复')}" if pick(row, m, "是否重复") else "",
        f"待核实:{pick(row, m, '是否待核实')}" if pick(row, m, "是否待核实") else "",
    ]
    return {
        "唯一编号": _gen_uid(platform, published, account, text),
        "采集时间": _parse_dt(pick(row, m, "采集时间")) or collected_default,
        "发布时间": published,
        "平台": platform,
        "平台组": platform_group(platform),
        "具体来源": pick(row, m, "具体来源") or pick(row, m, "来源类型"),
        "账号": account,
        "正文": text,
        "原始链接": pick(row, m, "原始链接"),
        "地区": region,
        "地区组": region_group,
        "省份": province,
        "城市": "",
        "IP属地": "",
        "语言": language,
        "是否少数民族语言": "是" if is_minority_lang(language) else "否",
        "总体态度": attitude,
        "具体问题类别": pick(row, m, "议题"),
        "点赞量": _to_int(pick(row, m, "点赞")),
        "评论量": _to_int(pick(row, m, "评论")),
        "转发量": _to_int(pick(row, m, "转发")),
        "是否重点": "否",
        "是否计入统计": counted_flag,
        "备注": "；".join(p for p in notes_parts if p),
    }


def normalize_comment(row, m, fname, kw_map, collected_default):
    platform = pick(row, m, "平台")
    text = pick(row, m, "评论正文") or pick(row, m, "正文") or pick(row, m, "标题")
    if not platform or not text:
        return None
    published = pick(row, m, "评论日期") or pick(row, m, "母帖日期") or pick(row, m, "发布时间")
    published = _parse_dt(published)
    account = pick(row, m, "评论昵称") or pick(row, m, "账号") or pick(row, m, "母帖账号")
    region = pick(row, m, "地区") or pick(row, m, "评论者地区")
    province = _infer_province(region, "", "", kw_map)
    mother_title = pick(row, m, "标题")
    notes_parts = [
        f"历史导入:{os.path.basename(fname)}",
        f"母帖:{mother_title}" if mother_title else "",
        f"信息量:{pick(row, m, '信息量')}" if pick(row, m, "信息量") else "",
        f"复核:{pick(row, m, '复核说明')}" if pick(row, m, "复核说明") else "",
        f"证据:{pick(row, m, '证据来源')}" if pick(row, m, "证据来源") else "",
    ]
    counted = pick(row, m, "是否计入")
    counted_flag = "否" if counted in ("否", "0", "不计入", "不纳入") else "是"
    return {
        "唯一编号": _gen_uid(platform, published, account, text),
        "采集时间": collected_default,
        "发布时间": published,
        "平台": platform,
        "平台组": platform_group(platform),
        "具体来源": pick(row, m, "母帖账号") or "公众评论明细",
        "账号": account,
        "正文": text,
        "原始链接": pick(row, m, "原始链接"),
        "地区": region,
        "地区组": "",
        "省份": province,
        "城市": "",
        "IP属地": "",
        "语言": pick(row, m, "语言") or "中文",
        "是否少数民族语言": "是" if is_minority_lang(pick(row, m, "语言") or "中文") else "否",
        "总体态度": pick(row, m, "态度") or "待核实",
        "具体问题类别": pick(row, m, "议题"),
        "点赞量": _to_int(pick(row, m, "点赞")),
        "评论量": 0,
        "转发量": 0,
        "是否重点": "否",
        "是否计入统计": counted_flag,
        "备注": "；".join(p for p in notes_parts if p),
    }


def normalize_template(row, m, fname, kw_map, collected_default):
    platform = pick(row, m, "平台")
    text = pick(row, m, "正文") or pick(row, m, "标题") or pick(row, m, "摘要")
    if not platform or not text:
        return None
    attitude = pick(row, m, "态度") or "待核实"
    if attitude == "排除记录":
        return None
    relevance = pick(row, m, "关联度")
    counted_flag = "否" if any(k in relevance for k in ("不计入", "无关", "排除")) else "是"
    published = _parse_dt(pick(row, m, "发布时间"))
    region = pick(row, m, "地区")
    province = _infer_province(region, "", "", kw_map)
    notes_parts = [
        f"历史导入:{os.path.basename(fname)}",
        f"数据编号:{pick(row, m, '数据编号')}" if pick(row, m, "数据编号") else "",
        f"来源类型:{pick(row, m, '来源类型')}" if pick(row, m, "来源类型") else "",
        f"检索关键词:{pick(row, m, '检索关键词')}" if pick(row, m, "检索关键词") else "",
        f"标题:{pick(row, m, '标题')}" if pick(row, m, "标题") else "",
    ]
    return {
        "唯一编号": _gen_uid(platform, published, pick(row, m, "账号"), text),
        "采集时间": _parse_dt(pick(row, m, "采集时间")) or collected_default,
        "发布时间": published,
        "平台": platform,
        "平台组": platform_group(platform),
        "具体来源": pick(row, m, "来源类型") or platform,
        "账号": pick(row, m, "账号"),
        "正文": text,
        "原始链接": pick(row, m, "原始链接"),
        "地区": region,
        "地区组": "",
        "省份": province,
        "城市": "",
        "IP属地": "",
        "语言": pick(row, m, "语言") or "中文",
        "是否少数民族语言": "是" if is_minority_lang(pick(row, m, "语言") or "中文") else "否",
        "总体态度": attitude,
        "具体问题类别": pick(row, m, "议题"),
        "点赞量": _to_int(pick(row, m, "点赞")),
        "评论量": _to_int(pick(row, m, "评论")),
        "转发量": _to_int(pick(row, m, "转发")),
        "是否重点": "否",
        "是否计入统计": counted_flag,
        "备注": "；".join(p for p in notes_parts if p),
    }


def normalize_nonsupport(row, m, fname, kw_map, collected_default):
    platform = pick(row, m, "平台")
    text = pick(row, m, "正文") or pick(row, m, "标题") or pick(row, m, "摘要")
    if not platform or not text:
        return None
    attitude = pick(row, m, "态度") or "待核实"
    if attitude == "排除记录":
        return None
    published = _parse_dt(pick(row, m, "发布时间"))
    region = pick(row, m, "地区")
    province = _infer_province(region, "", "", kw_map)
    counted = pick(row, m, "是否计入")
    counted_flag = "否" if any(k in counted for k in ("否", "0", "不计入", "不纳入", "排除")) else "是"
    notes_parts = [
        f"历史导入:{os.path.basename(fname)}",
        f"证据状态:{pick(row, m, '证据状态')}" if pick(row, m, "证据状态") else "",
        f"记录类型:{pick(row, m, '记录类型')}" if pick(row, m, "记录类型") else "",
        f"原表计数:{pick(row, m, '原表计数')}" if pick(row, m, "原表计数") else "",
        f"复核:{pick(row, m, '复核说明')}" if pick(row, m, "复核说明") else "",
        f"来源:{pick(row, m, '原始来源文件')}/{pick(row, m, '原始子表')}" if pick(row, m, "原始来源文件") else "",
        f"重复来源数:{pick(row, m, '重复来源数')}" if pick(row, m, "重复来源数") else "",
    ]
    return {
        "唯一编号": _gen_uid(platform, published, pick(row, m, "账号"), text),
        "采集时间": collected_default,
        "发布时间": published,
        "平台": platform,
        "平台组": platform_group(platform),
        "具体来源": pick(row, m, "记录类型") or platform,
        "账号": pick(row, m, "账号"),
        "正文": text,
        "原始链接": pick(row, m, "原始链接"),
        "地区": region,
        "地区组": "",
        "省份": province,
        "城市": "",
        "IP属地": "",
        "语言": "中文",
        "是否少数民族语言": "否",
        "总体态度": attitude,
        "具体问题类别": pick(row, m, "议题"),
        "点赞量": 0,
        "评论量": 0,
        "转发量": 0,
        "是否重点": "否",
        "是否计入统计": counted_flag,
        "备注": "；".join(p for p in notes_parts if p),
    }


def iter_source_rows(fname, kw_map, collected_default):
    import openpyxl

    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            hdr = find_detail_header(rows)
            if hdr is not None:
                m = col_index_map(hdr)
                if "平台" in m and ("正文" in m or "标题" in m):
                    hidx = next(i for i, r in enumerate(rows[:12]) if r is hdr)
                    for r in rows[hidx + 1:]:
                        rec = normalize_detail(r, m, fname, kw_map, collected_default)
                        if rec:
                            yield rec
                continue
            hdr = find_comment_header(rows)
            if hdr is not None:
                m = col_index_map(hdr)
                if "平台" in m and ("评论正文" in m or "正文" in m):
                    hidx = next(i for i, r in enumerate(rows[:12]) if r is hdr)
                    for r in rows[hidx + 1:]:
                        rec = normalize_comment(r, m, fname, kw_map, collected_default)
                        if rec:
                            yield rec
                continue
            hdr = find_template_header(rows)
            if hdr is not None:
                m = col_index_map(hdr)
                if "平台" in m and "正文" in m:
                    hidx = next(i for i, r in enumerate(rows[:12]) if r is hdr)
                    for r in rows[hidx + 1:]:
                        rec = normalize_template(r, m, fname, kw_map, collected_default)
                        if rec:
                            yield rec
                continue
            hdr = find_nonsupport_header(rows)
            if hdr is not None:
                m = col_index_map(hdr)
                if "平台" in m and "正文" in m:
                    hidx = next(i for i, r in enumerate(rows[:16]) if r is hdr)
                    for r in rows[hidx + 1:]:
                        rec = normalize_nonsupport(r, m, fname, kw_map, collected_default)
                        if rec:
                            yield rec
    finally:
        wb.close()


def main():
    parser = argparse.ArgumentParser(description="历史 Excel 全量导入舆情事实表")
    parser.add_argument("--xlsx", required=True, help="原子化工作簿路径")
    parser.add_argument("--source-dir", default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "02_data"))
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--no-clear", action="store_true", help="不清空事实表旧数据")
    parser.add_argument("--backup", action="store_true")
    args = parser.parse_args()

    xlsx = os.path.abspath(args.xlsx)
    source_dir = os.path.abspath(args.source_dir)
    if not os.path.exists(xlsx):
        print("[错误] 找不到工作簿:", xlsx)
        return 1
    if not os.path.isdir(source_dir):
        print("[错误] 找不到历史数据目录:", source_dir)
        return 1

    kw_map = _load_province_keywords(os.path.join(os.path.dirname(os.path.abspath(__file__)), "site_sources.json"))
    collected_default = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

    files = sorted(
        os.path.join(source_dir, f)
        for f in os.listdir(source_dir)
        if f.lower().endswith(".xlsx") and not f.startswith("~$")
    )
    print("历史数据目录:", source_dir)
    print("工作簿:", xlsx)
    print("待扫描文件:", len(files))

    seen = set()
    rows = []
    by_file = {}
    skipped = 0
    for f in files:
        before = len(seen)
        cnt = 0
        try:
            for rec in iter_source_rows(f, kw_map, collected_default):
                uid = rec["唯一编号"]
                if uid in seen:
                    skipped += 1
                    continue
                seen.add(uid)
                rows.append(rec)
                cnt += 1
        except Exception as e:
            print(f"[跳过文件] {os.path.basename(f)}: {e}")
            continue
        by_file[os.path.basename(f)] = cnt
        print(f"[{os.path.basename(f)}] 新增 {cnt} 条（文件内去重后）")

    print("\n唯一记录总数:", len(rows), "| 重复跳过:", skipped)
    print("按文件新增：")
    for k, v in sorted(by_file.items(), key=lambda x: -x[1]):
        print("  ", v, k)

    att = {}
    prov_missing = 0
    for r in rows:
        att[r["总体态度"]] = att.get(r["总体态度"], 0) + 1
        if not r["省份"]:
            prov_missing += 1
    print("态度分布:", dict(sorted(att.items(), key=lambda x: -x[1])))
    print("省份缺失:", prov_missing)

    if args.dry_run:
        print("\ndry-run：不写工作簿")
        return 0

    import openpyxl

    if args.backup:
        backup_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "backups")
        os.makedirs(backup_dir, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(backup_dir, f"{stamp}_{os.path.basename(xlsx)}")
        shutil.copy2(xlsx, backup_path)
        print("[备份]", backup_path)

    wb = openpyxl.load_workbook(xlsx)
    fact_ws = wb["舆情事实表"]
    fact_header = find_header_row(fact_ws, ("唯一编号", "正文", "是否计入统计"))
    if fact_header is None:
        print("[错误] 找不到舆情事实表表头")
        return 1
    fact_cols = column_index(fact_ws, fact_header, FACT_FIELDS)
    if not args.no_clear:
        fact_ws.delete_rows(fact_header + 1, fact_ws.max_row)
        print("[清空] 舆情事实表旧数据已删除（保留表头）")
        if "待清洗区" in wb.sheetnames:
            sws = wb["待清洗区"]
            sh = find_header_row(sws, ("清洗状态",))
            if sh:
                sws.delete_rows(sh + 1, sws.max_row)
                print("[清空] 待清洗区旧数据已删除（保留表头）")

    start = fact_ws.max_row + 1
    for ri, rec in enumerate(rows):
        excel_row = start + ri
        for field, ci in fact_cols.items():
            if ci is None:
                continue
            val = rec[field]
            if field in ("采集时间", "发布时间") and val:
                val = datetime.strptime(val, "%Y-%m-%dT%H:%M:%S")
            fact_ws.cell(row=excel_row, column=ci, value=val)

    tmp_path = xlsx + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, xlsx)
    print(f"[写入] 已追加 {len(rows)} 行到舆情事实表 -> {xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
