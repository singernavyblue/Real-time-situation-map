#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""clean_and_append.py 清洗入事实表

读取 inbox/ 的原始记录（collect.py 输出的 24 字段 JSON，或人工导出的 json/jsonl/csv），
按“唯一编号”去重后追加进原子化工作簿的「舆情事实表」；
缺必填字段的记录自动写入「待清洗区」，供人工复核。

用法：
  python clean_and_append.py --xlsx 路径/Excel数据库改造示例.xlsx
  python clean_and_append.py                          # 自动查找原子化工作簿
  ATOMIC_XLSX=路径 python clean_and_append.py
  python clean_and_append.py --inbox /tmp/inbox --dry-run
  python clean_and_append.py --no-move                # 处理完不移动 inbox 文件
  python clean_and_append.py --backup                 # 覆盖前先备份工作簿
"""
import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import config
from ingest import is_minority_lang, issue_category, platform_group

FACT_FIELDS = [
    "唯一编号", "采集时间", "发布时间", "平台", "平台组", "具体来源", "账号", "正文", "原始链接",
    "地区", "地区组", "省份", "城市", "IP属地", "语言", "是否少数民族语言", "总体态度", "具体问题类别",
    "点赞量", "评论量", "转发量", "是否重点", "是否计入统计", "备注",
]

STAGING_FIELDS = [
    "唯一编号", "采集时间", "平台", "账号", "正文", "地区", "省份", "语言", "总体态度", "清洗状态", "未通过原因/待办",
]

REQUIRED = ("平台", "正文", "总体态度")

FIELD_ALIASES = {
    "唯一编号": ("唯一编号", "uid", "编号"),
    "采集时间": ("采集时间", "collected_at", "采集日期"),
    "发布时间": ("发布时间", "published_at", "发布日期", "评论日期", "日期"),
    "平台": ("平台", "platform", "平台/网站", "平台或网站"),
    "平台组": ("平台组", "platform_group"),
    "具体来源": ("具体来源", "source", "来源"),
    "账号": ("账号", "account", "账号名称", "账号或栏目名称", "账号名称/发布单位"),
    "正文": ("正文", "text", "评论原文", "原文证据摘录", "原话", "内容", "评论观点"),
    "原始链接": ("原始链接", "url", "链接"),
    "地区": ("地区", "region", "涉及地区"),
    "地区组": ("地区组", "region_group"),
    "省份": ("省份", "province", "省"),
    "城市": ("城市", "city"),
    "IP属地": ("IP属地", "ip_location", "IP"),
    "语言": ("语言", "language", "原始语言"),
    "是否少数民族语言": ("是否少数民族语言", "is_minority"),
    "总体态度": ("总体态度", "attitude", "态度", "意见类型"),
    "具体问题类别": ("具体问题类别", "issue_category"),
    "点赞量": ("点赞量", "likes", "点赞"),
    "评论量": ("评论量", "comments", "评论数"),
    "转发量": ("转发量", "shares", "转发"),
    "是否重点": ("是否重点", "is_key", "重点舆情"),
    "是否计入统计": ("是否计入统计", "counted", "是否计入"),
    "备注": ("备注", "notes", "其他备注", "说明"),
}


def _clean(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\u3000", " ").replace("\xa0", " ")).strip()


def _get(rec, field):
    for k in FIELD_ALIASES[field]:
        if k in rec and rec[k] is not None and str(rec[k]).strip() != "":
            return rec[k]
    return ""


def _to_int(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = _clean(v).replace(",", "").replace("，", "")
    m = re.search(r"([\d.]+)\s*(万|w|W|k|K)?", s)
    if not m:
        return 0
    val = float(m.group(1))
    unit = m.group(2)
    if unit and unit.lower() == "万":
        val *= 10000
    elif unit and unit.lower() == "k":
        val *= 1000
    return int(val)


def _to_bool(v):
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    return str(v).strip() in ("是", "1", "TRUE", "True", "true", "计入", "已计入", "Y", "y")


def _parse_dt(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%dT%H:%M:%S")
    s = _clean(v)
    if not s:
        return ""
    m = re.search(r"(\d{4})[-/年.](\d{1,2})[-/月.](\d{1,2})(?:[ T日](\d{1,2})[:：点时](\d{1,2})?(?:[:：分](\d{1,2}))?)?", s)
    if not m:
        return ""
    try:
        dt = datetime(
            int(m.group(1)), int(m.group(2)), int(m.group(3)),
            int(m.group(4) or 0), int(m.group(5) or 0), int(m.group(6) or 0),
        )
        return dt.strftime("%Y-%m-%dT%H:%M:%S")
    except ValueError:
        return ""


def _dt_obj(v):
    if not v:
        return ""
    try:
        return datetime.strptime(v, "%Y-%m-%dT%H:%M:%S")
    except ValueError:
        return v


def _gen_uid(rec):
    key = "|".join([
        str(_get(rec, "平台") or ""),
        str(_get(rec, "发布时间") or ""),
        str(_get(rec, "账号") or ""),
        str(_get(rec, "正文") or ""),
    ])
    return "auto-" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:16]


def normalize_row(rec):
    row = {f: "" for f in FACT_FIELDS}
    row["唯一编号"] = _clean(_get(rec, "唯一编号")) or _gen_uid(rec)
    row["采集时间"] = _parse_dt(_get(rec, "采集时间")) or datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    row["发布时间"] = _parse_dt(_get(rec, "发布时间"))
    row["平台"] = _clean(_get(rec, "平台"))
    row["平台组"] = _clean(_get(rec, "平台组")) or platform_group(row["平台"])
    row["具体来源"] = _clean(_get(rec, "具体来源"))
    row["账号"] = _clean(_get(rec, "账号"))
    row["正文"] = _clean(_get(rec, "正文"))
    row["原始链接"] = _clean(_get(rec, "原始链接"))
    row["地区"] = _clean(_get(rec, "地区"))
    row["地区组"] = _clean(_get(rec, "地区组"))
    row["省份"] = _clean(_get(rec, "省份")) or row["地区"]
    row["城市"] = _clean(_get(rec, "城市"))
    row["IP属地"] = _clean(_get(rec, "IP属地"))
    row["语言"] = _clean(_get(rec, "语言")) or "中文"
    raw_min = _get(rec, "是否少数民族语言")
    if _to_bool(raw_min):
        row["是否少数民族语言"] = "是"
    elif raw_min and not _to_bool(raw_min) and _clean(raw_min) not in ("", "否", "0", "False", "false"):
        row["是否少数民族语言"] = "是"
    else:
        row["是否少数民族语言"] = "是" if is_minority_lang(row["语言"]) else "否"
    row["总体态度"] = _clean(_get(rec, "总体态度"))
    row["具体问题类别"] = _clean(_get(rec, "具体问题类别"))
    if row["具体问题类别"] in ("", "-", "无"):
        row["具体问题类别"] = issue_category(row["总体态度"])
    row["点赞量"] = _to_int(_get(rec, "点赞量"))
    row["评论量"] = _to_int(_get(rec, "评论量"))
    row["转发量"] = _to_int(_get(rec, "转发量"))
    row["是否重点"] = "是" if _to_bool(_get(rec, "是否重点")) else "否"
    raw_counted = _get(rec, "是否计入统计")
    row["是否计入统计"] = "是" if (raw_counted == "" or _to_bool(raw_counted)) else "否"
    row["备注"] = _clean(_get(rec, "备注"))
    return row


def validate_row(row):
    missing = [f for f in REQUIRED if not row[f]]
    return missing


def read_records(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".json":
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            for k in ("records", "items", "数据", "记录"):
                if isinstance(data.get(k), list):
                    return data[k]
            return [data]
        return []
    if ext == ".jsonl":
        out = []
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    out.append(json.loads(line))
        return out
    if ext == ".csv":
        out = []
        with open(path, encoding="utf-8-sig", newline="") as f:
            for r in csv.DictReader(f):
                out.append(dict(r))
        return out
    return []


def find_header_row(ws, keys):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        joined = "|".join(str(c) for c in row if c is not None)
        if all(k in joined for k in keys):
            return i
    return None


def column_index(ws, header_row, fields):
    header = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    idx = {}
    for i, h in enumerate(header, start=1):
        if h is not None and str(h).strip():
            idx.setdefault(str(h).strip(), i)
    return {f: idx.get(f) for f in fields}


def append_rows(ws, header_row, col_idx, rows):
    start = ws.max_row + 1
    for ri, row in enumerate(rows):
        excel_row = start + ri
        for field, ci in col_idx.items():
            if ci is None:
                continue
            val = row[field]
            if field in ("采集时间", "发布时间"):
                val = _dt_obj(val)
            ws.cell(row=excel_row, column=ci, value=val)
    return len(rows)


def detect_xlsx():
    env = os.environ.get("ATOMIC_XLSX", "").strip()
    if env and os.path.exists(env):
        return env
    roots = [
        os.path.abspath(os.path.join(config.BASE_DIR, "..", "..", "..", "outputs")),
        os.path.join(config.BASE_DIR, "docs"),
        config.BASE_DIR,
    ]
    for root in roots:
        if not os.path.isdir(root):
            continue
        for dirpath, dirs, files in os.walk(root):
            for fn in files:
                if fn.lower().endswith(".xlsx") and ("原子化" in fn or ("数据库" in fn and "示例" in fn)):
                    return os.path.join(dirpath, fn)
            depth = dirpath[len(root):].count(os.sep)
            if depth >= 2:
                dirs[:] = []
    return None


def main():
    parser = argparse.ArgumentParser(description="清洗 inbox/ 原始记录并追加进舆情事实表")
    parser.add_argument("--xlsx", default=None, help="原子化工作簿路径（默认 ATOMIC_XLSX 或自动查找）")
    parser.add_argument("--inbox", default=config.INBOX_DIR, help="原始记录目录（默认 inbox/）")
    parser.add_argument("--dry-run", action="store_true", help="只统计不写文件")
    parser.add_argument("--no-move", action="store_true", help="处理完不移动原始文件")
    parser.add_argument("--backup", action="store_true", help="覆盖工作簿前先备份")
    args = parser.parse_args()

    xlsx = args.xlsx or detect_xlsx()
    if not xlsx or not os.path.exists(xlsx):
        print("[错误] 找不到原子化工作簿，请用 --xlsx 指定或设置 ATOMIC_XLSX")
        return 1
    print("工作簿:", xlsx)
    print("输入目录:", args.inbox)

    import openpyxl

    wb = openpyxl.load_workbook(xlsx)
    if "舆情事实表" not in wb.sheetnames:
        print("[错误] 工作簿中没有“舆情事实表”")
        return 1
    fact_ws = wb["舆情事实表"]
    staging_ws = wb["待清洗区"] if "待清洗区" in wb.sheetnames else None

    fact_header = find_header_row(fact_ws, ("唯一编号", "正文", "是否计入统计"))
    if fact_header is None:
        print("[错误] 舆情事实表表头未找到（需含：唯一编号 / 正文 / 是否计入统计）")
        return 1
    fact_cols = column_index(fact_ws, fact_header, FACT_FIELDS)
    if fact_cols["唯一编号"] is None:
        print("[错误] 舆情事实表缺少“唯一编号”列")
        return 1

    existing_uids = set()
    for row in fact_ws.iter_rows(min_row=fact_header + 1, values_only=True):
        v = row[fact_cols["唯一编号"] - 1] if fact_cols["唯一编号"] - 1 < len(row) else None
        if v is not None and str(v).strip():
            existing_uids.add(str(v).strip())

    staging_header = None
    staging_cols = {}
    if staging_ws is not None:
        staging_header = find_header_row(staging_ws, ("清洗状态",))
        if staging_header:
            staging_cols = column_index(staging_ws, staging_header, STAGING_FIELDS)

    new_rows = []
    staging_rows = []
    seen = set(existing_uids)
    dup_count = 0
    invalid_count = 0
    file_count = 0

    os.makedirs(args.inbox, exist_ok=True)
    files = sorted(
        f for f in os.listdir(args.inbox)
        if not f.startswith(".") and f.lower().endswith((".json", ".jsonl", ".csv"))
    )
    for fname in files:
        path = os.path.join(args.inbox, fname)
        try:
            records = read_records(path)
        except Exception as e:
            print(f"[跳过] {fname} 解析失败: {e}")
            continue
        file_count += 1
        for raw in records:
            row = normalize_row(raw)
            missing = validate_row(row)
            if missing:
                invalid_count += 1
                reason = "缺" + "、".join(missing)
                staging_rows.append({
                    "唯一编号": row["唯一编号"],
                    "采集时间": row["采集时间"],
                    "平台": row["平台"],
                    "账号": row["账号"],
                    "正文": row["正文"],
                    "地区": row["地区"],
                    "省份": row["省份"],
                    "语言": row["语言"],
                    "总体态度": row["总体态度"],
                    "清洗状态": "待复核",
                    "未通过原因/待办": reason,
                })
                continue
            if row["唯一编号"] in seen:
                dup_count += 1
                continue
            seen.add(row["唯一编号"])
            new_rows.append(row)
        print(f"[文件] {fname}: {len(records)} 条")

    print(f"\n统计：新增 {len(new_rows)}，重复跳过 {dup_count}，缺字段 {invalid_count}，处理文件 {file_count}")
    if args.dry_run:
        print("dry-run：不写工作簿、不移动文件")
        return 0

    if new_rows:
        append_rows(fact_ws, fact_header, fact_cols, new_rows)
        print(f"[事实表] 追加 {len(new_rows)} 行（现有唯一编号 {len(existing_uids)} 个）")
    if staging_rows and staging_ws is not None and staging_header and staging_cols["清洗状态"]:
        append_rows(staging_ws, staging_header, staging_cols, staging_rows)
        print(f"[待清洗区] 追加 {len(staging_rows)} 行")

    if args.backup:
        backup_dir = os.path.join(config.DATA_DIR, "backups")
        os.makedirs(backup_dir, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(backup_dir, f"{stamp}_{os.path.basename(xlsx)}")
        shutil.copy2(xlsx, backup_path)
        print("[备份]", backup_path)

    tmp_path = xlsx + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, xlsx)
    print("[保存]", xlsx)

    if not args.no_move:
        processed_dir = os.path.join(args.inbox, "processed")
        os.makedirs(processed_dir, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d%H%M%S")
        moved = 0
        for fname in files:
            src = os.path.join(args.inbox, fname)
            if not os.path.exists(src):
                continue
            dst = os.path.join(processed_dir, f"{stamp}_{fname}")
            try:
                os.rename(src, dst)
                moved += 1
            except Exception as e:
                print(f"[警告] 移动 {fname} 失败: {e}")
        print(f"[归档] 已移动 {moved} 个文件到 {processed_dir}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
